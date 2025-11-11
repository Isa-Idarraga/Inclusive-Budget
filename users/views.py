from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse
from .models import User
from django import forms
from .decorators import role_required
from .password_manager import PasswordManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


# Formulario personalizado para crear/editar usuarios
class UserManagementForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class': 'form-control',
            'id': 'password-field'
        }),
        required=False,
        help_text="Deja en blanco para mantener la contraseña actual (solo al editar)"
    )

    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'last_name', 'role', 'is_active']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'role': forms.Select(attrs={'class': 'form-control'}),
        }


# Formulario para editar el perfil propio del usuario
class UserProfileForm(forms.ModelForm):
    current_password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class': 'form-control',
            'id': 'current-password'
        }),
        required=False,
        label="Contraseña actual",
        help_text="Requerida solo si deseas cambiar tu contraseña"
    )
    new_password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class': 'form-control',
            'id': 'new-password'
        }),
        required=False,
        label="Nueva contraseña",
        help_text="Deja en blanco si no deseas cambiar tu contraseña"
    )
    confirm_password = forms.CharField(
        widget=forms.PasswordInput(attrs={
            'class': 'form-control',
            'id': 'confirm-password'
        }),
        required=False,
        label="Confirmar nueva contraseña"
    )

    class Meta:
        model = User
        fields = ['username', 'email', 'first_name', 'last_name']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
        }
        labels = {
            'username': 'Usuario',
            'email': 'Correo electrónico',
            'first_name': 'Nombre',
            'last_name': 'Apellidos',
        }

    def clean(self):
        cleaned_data = super().clean()
        new_password = cleaned_data.get('new_password')
        confirm_password = cleaned_data.get('confirm_password')
        current_password = cleaned_data.get('current_password')

        # Si se intenta cambiar la contraseña
        if new_password or confirm_password:
            if not current_password:
                raise forms.ValidationError('Debes ingresar tu contraseña actual para cambiarla.')
            
            if new_password != confirm_password:
                raise forms.ValidationError('Las contraseñas nuevas no coinciden.')
            
            if len(new_password) < 6:
                raise forms.ValidationError('La nueva contraseña debe tener al menos 6 caracteres.')

        return cleaned_data


def logout_view(request):
    """Vista personalizada de logout que maneja POST y GET"""
    if request.user.is_authenticated:
        auth_logout(request)
        login_url = reverse('users:login')
        return redirect(f"{login_url}?logout=success")
    return redirect('users:login')


@login_required
def home(request):
    mods = []
    for key, label in [
        ("projects", "Proyectos"),
        ("catalog", "Materiales"),
        ("dashboard", "Dashboard"),
    ]:
        if request.user.can_see(key):
            mods.append({"key": key, "label": label})
    return render(request, "home.html", {"modules": mods})


@role_required(User.JEFE)
def user_list(request):
    """Lista todos los usuarios del sistema - Solo JEFE"""
    users_list = User.objects.all().order_by('-date_joined')
    
    # Obtener todas las contraseñas almacenadas
    all_passwords = PasswordManager.get_all_passwords()
    
    # Crear lista de usuarios con sus contraseñas
    users_with_passwords = []
    for user in users_list:
        user_data = {
            'user': user,
            'password': all_passwords.get(user.username, {}).get('password', None)
        }
        users_with_passwords.append(user_data)
    
    return render(request, 'users/user_list.html', {
        'users_with_passwords': users_with_passwords
    })


@role_required(User.JEFE)
def export_users_excel(request):
    """Exportar lista de usuarios a Excel - Solo JEFE"""
    
    # Crear el libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Definir encabezados
    headers = [
        "ID Usuario",
        "Nombre de Usuario",
        "Nombre Completo",
        "Email",
        "Rol",
        "Contraseña",
        "Fecha de Creación",
        "Estado",
        "Último Acceso"
    ]
    
    # Escribir encabezados con estilo
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Obtener todos los usuarios
    users = User.objects.all().order_by('-date_joined')
    
    # Obtener todas las contraseñas almacenadas
    all_passwords = PasswordManager.get_all_passwords()
    
    # Escribir datos de usuarios
    for row_num, user in enumerate(users, 2):
        # ID Usuario
        ws.cell(row=row_num, column=1, value=user.id)
        
        # Nombre de Usuario
        ws.cell(row=row_num, column=2, value=user.username)
        
        # Nombre Completo
        full_name = user.get_full_name() or "-"
        ws.cell(row=row_num, column=3, value=full_name)
        
        # Email
        ws.cell(row=row_num, column=4, value=user.email or "-")
        
        # Rol
        ws.cell(row=row_num, column=5, value=user.get_role_display())
        
        # Contraseña (solo para COMERCIAL y CONSTRUCTOR)
        if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
            password = all_passwords.get(user.username, {}).get('password', 'No disponible')
            ws.cell(row=row_num, column=6, value=password)
        else:
            ws.cell(row=row_num, column=6, value="Oculta (JEFE)")
        
        # Fecha de Creación
        fecha_creacion = user.date_joined.strftime("%d/%m/%Y %H:%M") if user.date_joined else "-"
        ws.cell(row=row_num, column=7, value=fecha_creacion)
        
        # Estado (Activo/Inactivo)
        estado = "Activo" if user.is_active else "Inactivo"
        ws.cell(row=row_num, column=8, value=estado)
        
        # Último Acceso
        ultimo_acceso = user.last_login.strftime("%d/%m/%Y %H:%M") if user.last_login else "Nunca"
        ws.cell(row=row_num, column=9, value=ultimo_acceso)
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 12,  # ID Usuario
        'B': 20,  # Nombre de Usuario
        'C': 25,  # Nombre Completo
        'D': 30,  # Email
        'E': 20,  # Rol
        'F': 20,  # Contraseña
        'G': 20,  # Fecha de Creación
        'H': 15,  # Estado
        'I': 20,  # Último Acceso
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Preparar respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha actual
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    filename = f"Usuarios_{fecha_actual}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response


@role_required(User.JEFE)
def user_create(request):
    """Crear un nuevo usuario - Solo JEFE"""
    if request.method == 'POST':
        form = UserManagementForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('password')

            if password:
                user.set_password(password)
                # Guardar contraseña visible si NO es JEFE
                if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
                    PasswordManager.save_password(user.username, password, user.role)
            else:
                default_password = 'changeme123'
                user.set_password(default_password)  # Password por defecto
                # Guardar contraseña por defecto si NO es JEFE
                if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
                    PasswordManager.save_password(user.username, default_password, user.role)

            user.save()
            messages.success(request, f'✅ Usuario "{user.username}" creado exitosamente.')
            return redirect('users:user_list')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = UserManagementForm()

    return render(request, 'users/user_form.html', {'form': form, 'mode': 'create'})


@role_required(User.JEFE)
def user_update(request, user_id):
    """Editar un usuario existente - Solo JEFE"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = UserManagementForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('password')

            if password:
                user.set_password(password)
                # Actualizar contraseña visible si NO es JEFE
                if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
                    PasswordManager.save_password(user.username, password, user.role)
                    messages.success(request, f'✅ Usuario "{user.username}" actualizado exitosamente. Contraseña guardada.')
                else:
                    messages.success(request, f'✅ Usuario "{user.username}" actualizado exitosamente.')
            else:
                messages.success(request, f'✅ Usuario "{user.username}" actualizado exitosamente.')

            user.save()
            return redirect('users:user_list')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = UserManagementForm(instance=user)

    return render(request, 'users/user_form.html', {
        'form': form,
        'mode': 'update',
        'user_obj': user
    })


@role_required(User.JEFE)
def user_delete(request, user_id):
    """Eliminar un usuario - Solo JEFE"""
    user = get_object_or_404(User, id=user_id)

    # No permitir que se elimine a sí mismo
    if user.id == request.user.id:
        messages.error(request, '❌ No puedes eliminarte a ti mismo.')
        return redirect('users:user_list')

    if request.method == 'POST':
        username = user.username
        # Eliminar contraseña del archivo antes de eliminar usuario
        PasswordManager.delete_password(username)
        user.delete()
        messages.success(request, f'✅ Usuario "{username}" eliminado exitosamente.')
        return redirect('users:user_list')

    return render(request, 'users/user_confirm_delete.html', {'user_obj': user})


@role_required(User.JEFE)
def user_reset_password(request, user_id):
    """Resetear contraseña de un usuario - Solo JEFE"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        
        if not new_password:
            messages.error(request, '❌ Debes ingresar una contraseña.')
            return redirect('users:user_list')
        
        if len(new_password) < 4:
            messages.error(request, '❌ La contraseña debe tener al menos 4 caracteres.')
            return redirect('users:user_list')
        
        # Cambiar contraseña
        user.set_password(new_password)
        user.save()
        
        # Guardar contraseña visible si NO es JEFE
        if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
            PasswordManager.save_password(user.username, new_password, user.role)
            messages.success(request, f'🔑 Contraseña de "{user.username}" reseteada a: {new_password}')
        else:
            messages.success(request, f'🔑 Contraseña de "{user.username}" reseteada exitosamente.')
        
        return redirect('users:user_list')
    
    return render(request, 'users/user_reset_password.html', {'user_obj': user})


@login_required
def user_profile(request):
    """Vista para que cualquier usuario edite su propio perfil"""
    user = request.user

    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=user)
        if form.is_valid():
            # Verificar si se está intentando cambiar la contraseña
            new_password = form.cleaned_data.get('new_password')
            current_password = form.cleaned_data.get('current_password')

            if new_password:
                # Verificar que la contraseña actual sea correcta
                if not user.check_password(current_password):
                    messages.error(request, '❌ La contraseña actual es incorrecta.')
                    return render(request, 'users/user_profile.html', {
                        'form': form,
                        'user_obj': user
                    })

                # Cambiar la contraseña
                user.set_password(new_password)
                user.save()
                
                # GUARDAR CONTRASEÑA VISIBLE SI NO ES JEFE
                if user.role in ['COMERCIAL', 'CONSTRUCTOR']:
                    PasswordManager.save_password(user.username, new_password, user.role)
                
                # Cerrar sesión por seguridad y redirigir al login
                auth_logout(request)
                messages.success(request, '🔐 Contraseña actualizada exitosamente. Por favor, inicia sesión nuevamente con tu nueva contraseña.')
                return redirect('users:login')
            
            # Guardar los demás cambios (sin cambio de contraseña)
            form.save()
            messages.success(request, '✅ Perfil actualizado exitosamente.')
            return redirect('users:profile')
        else:
            messages.error(request, '❌ Por favor corrige los errores en el formulario.')
    else:
        form = UserProfileForm(instance=user)

    return render(request, 'users/user_profile.html', {
        'form': form,
        'user_obj': user
    })
