from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
import io
from django.core.paginator import Paginator
from django.db.models import Q, Max, Sum, F, Count
from .models import Project, Worker, Role, BudgetSection, BudgetItem, ProjectBudgetItem
from django.db.models import Q, Max, Sum, F, DecimalField, ExpressionWrapper
from catalog.models import Material, Supplier, MaterialSupplier
from django.utils import timezone
from zoneinfo import ZoneInfo
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
from .models import Project, Worker, Role, BudgetSection, BudgetItem, ProjectBudgetItem, ConsumoMaterial, ProyectoMaterial
from .forms import ProjectForm, WorkerForm, RoleForm, ConsumoMaterialForm, DetailedProjectForm, BudgetSectionForm, BudgetManagementForm, BudgetItemCreateForm, BudgetItemEditForm
import json
from django.urls import reverse
from .models import Project, EntradaMaterial, ConsumoMaterial
from .forms import EntradaMaterialForm
from users.decorators import role_required, project_owner_or_jefe_required
from users.models import User
from django.core.exceptions import PermissionDenied
from decimal import Decimal
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, ExpressionWrapper, FloatField
from django.shortcuts import render, get_object_or_404
from projects.models import Project, BudgetSection, BudgetItem, ConsumoMaterial
from django.http import HttpResponse
from .utils import get_etapas_con_avance

@login_required
def budget_progress_report(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    # Usar las secciones específicas del proyecto si existen, sino usar las plantillas globales
    secciones = BudgetSection.objects.filter(project=project)
    if not secciones.exists():
        secciones = BudgetSection.objects.filter(project__isnull=True)

    reporte = []
    for seccion in secciones:
        items = BudgetItem.objects.filter(section=seccion)
        presupuesto = sum(item.costo_unitario * item.cantidad for item in items)

        consumos = ConsumoMaterial.objects.filter(etapa=seccion)
        gasto = sum(c.cantidad * c.material.unit_cost for c in consumos)

        porcentaje = round((gasto / presupuesto) * 100, 2) if presupuesto > 0 else 0

        if gasto == 0:
            estado = "pendiente"
            color = "secondary"
            alerta = None
        elif porcentaje < 80:
            estado = "ok"
            color = "success"
            alerta = None
        elif porcentaje <= 100:
            estado = "medio"
            color = "warning"
            alerta = None
        else:
            estado = "sobrecosto"
            color = "danger"
            alerta = f"+{round(porcentaje - 100, 2)}% sobre presupuesto"

        reporte.append({
            "seccion": seccion,
            "presupuesto": presupuesto,
            "gastado": gasto,
            "porcentaje": porcentaje,
            "estado": estado,
            "color": color,
            "alerta": alerta
        })

    if "export" in request.GET:
        return export_reporte_excel(project, reporte)

    return render(request, "projects/budget_progress_report.html", {
        "project": project,
        "reporte": reporte,
    })

def export_reporte_excel(project, reporte):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avance por Etapa"

    headers = ["Etapa", "Presupuesto", "Gastado", "% Ejecutado", "Estado", "Alerta"]
    ws.append(headers)

    for r in reporte:
        ws.append([
            r["seccion"].name,
            r["presupuesto"],
            r["gastado"],
            r["porcentaje"],
            r["estado"],
            r["alerta"] or ""
        ])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_etapas_{project.name}.xlsx"'
    return response


# Helper function para obtener hora colombiana
def get_colombia_time():
    """Obtiene la hora actual en zona horaria de Colombia"""
    colombia_tz = ZoneInfo("America/Bogota")
    return timezone.now().astimezone(colombia_tz)

# Función para registrar entrada de material al inventario del proyecto
@project_owner_or_jefe_required
def registrar_entrada_material(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == "POST":
        form = EntradaMaterialForm(request.POST)
        if form.is_valid():
            entrada = form.save(commit=False)
            entrada.proyecto = project
            entrada.save()

            # Verificar si hay un consumo pendiente en la sesión (RF17D)
            consumo_pendiente = request.session.get('consumo_pendiente')

            if consumo_pendiente:
                try:
                    # Intentar registrar el consumo pendiente automáticamente
                    from .models import ConsumoMaterial
                    from catalog.models import Material

                    material = Material.objects.get(id=consumo_pendiente['material_id'])

                    # Crear el consumo con los datos guardados
                    consumo = ConsumoMaterial(
                        proyecto=project,
                        material=material,
                        cantidad_consumida=consumo_pendiente['cantidad_consumida'],
                        fecha_consumo=consumo_pendiente['fecha_consumo'],
                        componente_actividad=consumo_pendiente['componente_actividad'],
                        responsable=consumo_pendiente['responsable'],
                        observaciones=consumo_pendiente.get('observaciones', ''),
                        registrado_por=request.user
                    )

                    # Intentar guardar (validará stock nuevamente)
                    consumo.save()

                    # Limpiar la sesión
                    del request.session['consumo_pendiente']

                    messages.success(
                        request,
                        f'✅ Compra registrada exitosamente. '
                        f'✅ Consumo registrado automáticamente: {consumo.cantidad_consumida} {material.unit.symbol} '
                        f'de {material.name} para {consumo.componente_actividad}'
                    )

                except Exception as e:
                    # Si falla el registro del consumo, solo mostrar advertencia
                    messages.warning(
                        request,
                        f'✅ Compra registrada exitosamente. '
                        f'⚠️ No se pudo registrar el consumo automáticamente: {str(e)}. '
                        f'Por favor, regístralo manualmente.'
                    )
                    # Limpiar la sesión de todos modos
                    if 'consumo_pendiente' in request.session:
                        del request.session['consumo_pendiente']
            else:
                # No hay consumo pendiente, solo mensaje de compra exitosa
                messages.success(request, f'✅ Compra de {entrada.material.name} registrada exitosamente.')

            return redirect("projects:project_board", project_id=project.id)
    else:
        form = EntradaMaterialForm()

        # Verificar si hay un consumo pendiente para mostrar un aviso
        consumo_pendiente = request.session.get('consumo_pendiente')
        if consumo_pendiente:
            try:
                from catalog.models import Material
                material = Material.objects.get(id=consumo_pendiente['material_id'])
                messages.info(
                    request,
                    f'ℹ️ Después de registrar esta compra, se agregará automáticamente el consumo de '
                    f'{consumo_pendiente["cantidad_consumida"]} {material.unit.symbol} de {material.name}.'
                )
            except:
                pass

    return render(request, "projects/registrar_entrada_material.html", {"form": form, "project": project})


@login_required
def search_materials(request):
    """Devuelve materiales filtrados por nombre o código para el buscador dinámico."""
    term = request.GET.get("q", "").strip()

    if len(term) < 2:
        return JsonResponse({
            "results": [],
            "count": 0,
            "requires_more": True,
        })

    materials = (
        Material.objects.filter(
            Q(name__icontains=term) | Q(sku__icontains=term)
        )
        .select_related("unit")
        .order_by("name")[:20]
    )

    results = [
        {
            "id": material.id,
            "sku": material.sku,
            "name": material.name,
            "unit": material.unit.symbol,
            "unit_name": material.unit.name,
        }
        for material in materials
    ]

    return JsonResponse({
        "results": results,
        "count": len(results),
        "query": term,
    })


@login_required
def material_suppliers(request):
    """Lista proveedores disponibles para un material específico."""
    material_id = request.GET.get("material_id")

    if not material_id:
        return JsonResponse({"results": [], "message": "material_id requerido"}, status=400)

    suppliers = (
        MaterialSupplier.objects.filter(material_id=material_id)
        .select_related("supplier")
        .order_by("supplier__name")
    )

    results = [
        {
            "id": item.supplier.id,
            "name": item.supplier.name,
            "price": float(item.price) if item.price is not None else None,
            "preferred": item.preferred,
        }
        for item in suppliers
    ]

    return JsonResponse({
        "results": results,
        "count": len(results),
    })

@role_required(User.CONSTRUCTOR, User.JEFE)
def project_list(request):
    """
    Vista para listar proyectos
    - COMERCIAL: Ve todos los proyectos (solo lectura, info básica)
    - CONSTRUCTOR: Ve todos los proyectos (solo edita los suyos)
    - JEFE: Ve y puede editar todos los proyectos
    
    FUNCIONALIDAD DE ORDENAMIENTO:
    - Sin filtro de estado: Agrupa por estado (En Proceso → Futuros → Terminados)
    - "todos" (Todos los estados): Lista cronológica con ordenamiento toggleable
    """
    # Obtener parámetros de búsqueda y ordenamiento
    search_query = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")  # Vacío por defecto, "todos" para cronológico
    trabajadores_filter = request.GET.get("trabajadores", "")
    creador_filter = request.GET.get("creador", "")
    fecha_desde_filter = request.GET.get("fecha_desde", "")
    fecha_hasta_filter = request.GET.get("fecha_hasta", "")
    ubicacion_filter = request.GET.get("ubicacion", "")
    presupuesto_min_filter = request.GET.get("presupuesto_min", "")
    presupuesto_max_filter = request.GET.get("presupuesto_max", "")
    order_by = request.GET.get("order", "desc")  # desc = descendente (más reciente primero), asc = ascendente

    # TODOS los usuarios ven TODOS los proyectos
    projects = Project.objects.all()

    # Filtro por búsqueda - Buscar en nombre, descripción y ubicación
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query)  # Buscar en nombre
            | Q(description__icontains=search_query)  # Buscar en descripción
            | Q(location_address__icontains=search_query)  # Buscar en dirección
        )

    # Filtro por estado
    # Si status_filter es "todos", no filtramos, solo ordenamos cronológicamente
    if status_filter and status_filter != "todos":
        projects = projects.filter(estado=status_filter)

    # Filtro por número de trabajadores
    if trabajadores_filter:
        if trabajadores_filter == "0":
            projects = projects.filter(workers__isnull=True)
        elif trabajadores_filter == "1-3":
            projects = projects.annotate(worker_count=Count('workers')).filter(worker_count__gte=1, worker_count__lte=3)
        elif trabajadores_filter == "4-6":
            projects = projects.annotate(worker_count=Count('workers')).filter(worker_count__gte=4, worker_count__lte=6)
        elif trabajadores_filter == "7+":
            projects = projects.annotate(worker_count=Count('workers')).filter(worker_count__gte=7)

    # Filtro por creador
    if creador_filter:
        try:
            creador_id = int(creador_filter)
            projects = projects.filter(creado_por_id=creador_id)
        except ValueError:
            pass

    # Filtro por fecha desde
    if fecha_desde_filter:
        projects = projects.filter(fecha_creacion__gte=fecha_desde_filter)

    # Filtro por fecha hasta
    if fecha_hasta_filter:
        projects = projects.filter(fecha_creacion__lte=fecha_hasta_filter)

    # Filtro por ubicación (búsqueda de texto en dirección)
    if ubicacion_filter:
        projects = projects.filter(location_address__icontains=ubicacion_filter)
        print(f"DEBUG: Filtro ubicación aplicado: {ubicacion_filter}")

    # Filtro por presupuesto mínimo
    if presupuesto_min_filter:
        try:
            # Limpiar puntos y convertir a float
            presupuesto_min_clean = presupuesto_min_filter.replace('.', '').replace(',', '')
            presupuesto_min = float(presupuesto_min_clean)
            if presupuesto_min > 0:
                projects = projects.filter(presupuesto__gte=presupuesto_min)
                print(f"DEBUG: Filtro presupuesto mínimo aplicado: {presupuesto_min}")
        except (ValueError, AttributeError) as e:
            print(f"DEBUG: Error en filtro presupuesto mínimo: {e}")

    # Filtro por presupuesto máximo
    if presupuesto_max_filter:
        try:
            # Limpiar puntos y convertir a float
            presupuesto_max_clean = presupuesto_max_filter.replace('.', '').replace(',', '')
            presupuesto_max = float(presupuesto_max_clean)
            if presupuesto_max > 0:
                projects = projects.filter(presupuesto__lte=presupuesto_max)
                print(f"DEBUG: Filtro presupuesto máximo aplicado: {presupuesto_max}")
        except (ValueError, AttributeError) as e:
            print(f"DEBUG: Error en filtro presupuesto máximo: {e}")

    # Determinar cómo organizar los proyectos
    show_chronological = (status_filter == "todos")
    
    if show_chronological:
        # MODO CRONOLÓGICO: Todos los proyectos ordenados por fecha
        if order_by == "asc":
            projects_chronological = projects.order_by('fecha_creacion')  # Más antiguos primero
        else:
            projects_chronological = projects.order_by('-fecha_creacion')  # Más recientes primero (default)
        
        context = {
            "show_chronological": True,
            "projects_chronological": projects_chronological,
            "order_by": order_by,
            "search_query": search_query,
            "status_filter": status_filter,
            "trabajadores_filter": trabajadores_filter,
            "creador_filter": creador_filter,
            "creador_nombre": None,
            "fecha_desde_filter": fecha_desde_filter,
            "fecha_hasta_filter": fecha_hasta_filter,
            "ubicacion_filter": ubicacion_filter,
            "presupuesto_min_filter": presupuesto_min_filter,
            "presupuesto_max_filter": presupuesto_max_filter,
            "creadores": User.objects.filter(
                id__in=Project.objects.values_list('creado_por_id', flat=True).distinct()
            ).order_by('first_name', 'last_name'),
        }
    else:
        # MODO AGRUPADO POR ESTADO (default)
        # Separar proyectos por estado sin ordenar por fecha
        projects_en_proceso = projects.filter(estado='en_proceso')
        projects_terminados = projects.filter(estado='terminado')
        projects_futuros = projects.filter(estado='futuro')

        # Obtener lista de creadores para el filtro
        creadores = User.objects.filter(
            id__in=Project.objects.values_list('creado_por_id', flat=True).distinct()
        ).order_by('first_name', 'last_name')

        # Obtener el nombre del creador seleccionado para mostrarlo en el filtro activo
        creador_nombre = None
        if creador_filter:
            try:
                creador_id = int(creador_filter)
                creador_obj = User.objects.filter(id=creador_id).first()
                if creador_obj:
                    creador_nombre = f"{creador_obj.first_name} {creador_obj.last_name}".strip()
                    if not creador_nombre:  # Si no tiene nombre, usar username
                        creador_nombre = creador_obj.username
            except ValueError:
                pass

        context = {
            "show_chronological": False,
            "projects_en_proceso": projects_en_proceso,
            "projects_terminados": projects_terminados,
            "projects_futuros": projects_futuros,
            "search_query": search_query,
            "status_filter": status_filter,
            "trabajadores_filter": trabajadores_filter,
            "creador_filter": creador_filter,
            "creador_nombre": creador_nombre,
            "fecha_desde_filter": fecha_desde_filter,
            "fecha_hasta_filter": fecha_hasta_filter,
            "ubicacion_filter": ubicacion_filter,
            "presupuesto_min_filter": presupuesto_min_filter,
            "presupuesto_max_filter": presupuesto_max_filter,
            "creadores": creadores,
        }

    return render(request, "projects/project_list.html", context)


@login_required
def project_history(request):
    """
    Vista para mostrar el historial cronológico de proyectos
    Muestra todos los proyectos activos y finalizados ordenados cronológicamente
    """
    # Obtener parámetros de búsqueda y ordenamiento
    search_query = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")
    order_by = request.GET.get("order", "oldest")  # oldest o newest
    
    # Obtener todos los proyectos activos y finalizados
    projects = Project.objects.filter(estado__in=['en_proceso', 'terminado', 'futuro'])
    
    # Filtro por búsqueda
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(location_address__icontains=search_query)
        )
    
    # Filtro por estado
    if status_filter:
        projects = projects.filter(estado=status_filter)
    
    # Ordenar cronológicamente según la opción seleccionada
    if order_by == "newest":
        projects = projects.order_by('-fecha_creacion')  # Más recientes primero
    else:
        projects = projects.order_by('fecha_creacion')   # Más antiguos primero (default)
    
    # Separar por estado manteniendo el orden cronológico
    projects_en_proceso = projects.filter(estado='en_proceso')
    projects_terminados = projects.filter(estado='terminado')
    projects_futuros = projects.filter(estado='futuro')
    
    context = {
        "projects_en_proceso": projects_en_proceso,
        "projects_terminados": projects_terminados,
        "projects_futuros": projects_futuros,
        "search_query": search_query,
        "status_filter": status_filter,
        "order_by": order_by,
        "is_history_view": True,  # Flag para identificar que es la vista de historial
    }
    
    return render(request, "projects/project_history.html", context)


@login_required
@login_required
def project_board(request, project_id):
    """
    Vista tablero del proyecto: nombre, presupuesto, calendario, botones y listado de compras.
    - JEFE: Siempre accede al tablero (todos los proyectos)
    - CONSTRUCTOR: Solo accede al tablero si él creó el proyecto
    """
    # Obtener proyecto con sus entradas relacionadas a material y proveedor
    project = get_object_or_404(
        Project.objects.prefetch_related(
            'entradas__material', 
            'entradas__proveedor'
        ),
        id=project_id
    )

    # Verificar permisos: JEFE siempre puede, otros solo si crearon el proyecto
    if request.user.role != User.JEFE and not request.user.is_superuser:
        if project.creado_por != request.user:
            # Si no es JEFE y no creó el proyecto, redirigir a detalles
            return redirect('projects:project_detail', project_id=project.id)

    # Entradas de materiales del proyecto
    entradas_raw = project.entradas.all().order_by('material__name', '-fecha_ingreso')

    # Pre-cargar precios por proveedor para evitar consultas repetidas
    material_ids = {entrada.material_id for entrada in entradas_raw}
    supplier_ids = {entrada.proveedor_id for entrada in entradas_raw if entrada.proveedor_id}
    supplier_price_map = {}
    if material_ids and supplier_ids:
        supplier_price_map = {
            (mp.material_id, mp.supplier_id): mp.price
            for mp in MaterialSupplier.objects.filter(
                material_id__in=material_ids,
                supplier_id__in=supplier_ids
            )
        }

    # Importar el modelo de consumos
    from .models import ConsumoMaterial, ProyectoMaterial

    # Agrupar por material
    from collections import defaultdict
    materiales_agrupados = defaultdict(lambda: {
        'material': None,
        'entradas': [],
        'cantidad_total': 0,
        'stock_proyecto': 0,
        'consumos': [],
        'cantidad_consumida': 0,
        'costo_total': Decimal('0')
    })

    for entrada in entradas_raw:
        material_id = entrada.material.id

        if materiales_agrupados[material_id]['material'] is None:
            materiales_agrupados[material_id]['material'] = entrada.material

            # Obtener consumos de este material en este proyecto
            consumos = ConsumoMaterial.objects.filter(
                proyecto=project,
                material=entrada.material
            ).select_related('registrado_por').order_by('-fecha_consumo')

            materiales_agrupados[material_id]['consumos'] = list(consumos)
            materiales_agrupados[material_id]['cantidad_consumida'] = sum(
                c.cantidad_consumida for c in consumos
            )

        materiales_agrupados[material_id]['entradas'].append(entrada)
        materiales_agrupados[material_id]['cantidad_total'] += entrada.cantidad

        # Determinar precio por unidad
        unit_price = None
        if entrada.proveedor_id:
            unit_price = supplier_price_map.get((entrada.material_id, entrada.proveedor_id))
        if unit_price is None and entrada.material.unit_cost is not None:
            unit_price = entrada.material.unit_cost

        entrada.unit_price_display = unit_price
        if unit_price is not None:
            total_price = unit_price * entrada.cantidad
            entrada.total_price_display = total_price
            materiales_agrupados[material_id]['costo_total'] += total_price
        else:
            entrada.total_price_display = None

    # Calcular el stock correcto para cada material después de sumar todas las entradas
    for material_id, data in materiales_agrupados.items():
        # Stock = Total comprado - Total consumido
        data['stock_proyecto'] = data['cantidad_total'] - data['cantidad_consumida']

    # Convertir a lista para el template
    compras = list(materiales_agrupados.values())

    context = {
        "project": project,
        "compras": compras,
        "details_url": reverse("projects:project_detail", kwargs={"project_id": project.id}),
        "add_purchases_url": reverse("projects:registrar_entrada_material", kwargs={"project_id": project.id}),
        "charts_url": f"/proyectos/{project.id}/graficos/",
    }

    return render(request, "projects/project_board.html", context)


@login_required
def project_create(request):
    """
    Vista para crear un nuevo proyecto
    - COMERCIAL: Solo puede crear (después no podrá verlo en la lista)
    - CONSTRUCTOR: Redirige al formulario detallado
    - JEFE: Redirige al formulario detallado
    """
    # Redirigir JEFE y CONSTRUCTOR al formulario detallado
    if request.user.role in [User.JEFE, User.CONSTRUCTOR]:
        return redirect("projects:detailed_project_create")
    
    # Solo COMERCIAL usa el formulario simple
    workers = Worker.objects.all()
    if request.method == "POST":
        form = ProjectForm(request.POST, request.FILES)
        selected_workers = request.POST.getlist("workers")
        if form.is_valid():
            try:
                project = form.save(commit=False)
                project.creado_por = request.user

                from decimal import Decimal
                from .utils import create_default_budget_sections
                project.built_area = project.area_construida_total or Decimal("0")
                project.exterior_area = project.area_exterior_intervenir or Decimal("0")
                project.columns_count = project.columns_count or 0
                project.walls_area = project.walls_area or Decimal("0")
                project.windows_area = project.windows_area or Decimal("0")
                project.doors_count = project.doors_count or 0
                project.doors_height = Decimal("2.1")

                project.save()
                # ✅ Crear automáticamente las secciones del presupuesto
                # create_default_budget_sections(project)
                
                project.calculate_legacy_fields()
                project.presupuesto = project.calculate_final_budget()
                project.save()
                

                if selected_workers:
                    project.workers.set(selected_workers)

                messages.success(
                    request,
                    f'✅ Proyecto "{project.name}" creado exitosamente! Presupuesto estimado: ${project.presupuesto:,.0f}',
                )

                # Redirigir según el rol
                if request.user.role == User.COMERCIAL:
                    # COMERCIAL no puede ver el detalle, mostrar mensaje y redirigir a crear otro
                    messages.info(request, "Proyecto creado. Puedes crear otro presupuesto.")
                    return redirect("projects:project_create")
                else:
                    return redirect("projects:project_detail", project_id=project.id)

            except Exception as e:
                messages.error(request, f"❌ Error al crear el proyecto: {str(e)}")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        form = ProjectForm()

    return render(
        request,
        "projects/project_form.html",
        {
            "form": form,
            "workers": workers,
            "no_workers": not workers.exists(),
        },
    )

@login_required
def project_detail(request, project_id):
    """
    Vista para mostrar el detalle de un proyecto
    Accesible para todos los usuarios autenticados
    """
    # Obtener proyecto sin restricción de creador
    project = get_object_or_404(Project, id=project_id)

    # Entradas de materiales del proyecto
    compras = project.entradas.select_related("material", "proveedor").all()

    # Agregar stock por proyecto a cada entrada
    for compra in compras:
        pm = compra.material.proyectos.filter(proyecto=project).first()
        compra.stock_proyecto = pm.stock_proyecto if pm else 0

    # Calcular presupuesto estimado usando los datos del proyecto
    project.calculate_legacy_fields()
    estimated_budget = project.calculate_final_budget()

    # Obtener avance por etapa del presupuesto
    etapas_con_avance = get_etapas_con_avance(project)
    
    avance_por_etapas = get_etapas_con_avance(project)
    # print("DEBUG avance_por_etapas:", avance_por_etapas)


    context = {
        'project': project,
        'compras': compras,
        'estimated_budget': estimated_budget,
        'etapas_con_avance': etapas_con_avance,
    }

    return render(request, "projects/project_detail.html", context)


@login_required
def detalle_etapa_consumos(request, etapa_id):
    etapa = get_object_or_404(BudgetSection, id=etapa_id)

    # Obtener consumos relacionados con la etapa, pero limitar por proyecto cuando sea posible
    consumos_qs = etapa.consumos_materiales.select_related("material", "registrado_por").all()

    # Si se proporciona ?project=<id> en la URL, lo usamos para filtrar (viene desde project_detail)
    project_filter = request.GET.get('project')
    if project_filter:
        try:
            pid = int(project_filter)
            consumos_qs = consumos_qs.filter(proyecto_id=pid)
        except (ValueError, TypeError):
            # Ignorar filtro inválido
            pass
    # Si la etapa pertenece a un proyecto concreto, filtrar por ese proyecto
    elif getattr(etapa, 'project_id', None):
        consumos_qs = consumos_qs.filter(proyecto_id=etapa.project_id)

    consumos = consumos_qs

    consumos_con_totales = []
    for c in consumos:
        total = (c.cantidad_consumida or 0) * (c.material.unit_cost or 0)
        consumos_con_totales.append({
            "fecha": c.fecha_registro,
            "material": c.material.name,
            "cantidad": c.cantidad_consumida,
            "unit_cost": c.material.unit_cost,
            "total": total,
            "responsable": c.registrado_por.get_full_name() if c.registrado_por else "—",
        })

    return render(request, "projects/detalle_etapa_consumos.html", {
        "etapa": etapa,
        "consumos": consumos_con_totales
    })


@login_required
def exportar_avance_etapas_excel(request, project_id):
    """
    Exporta el avance por etapas del presupuesto a un archivo Excel.
    """
    project = get_object_or_404(Project, id=project_id)
    etapas_con_avance = get_etapas_con_avance(project)

    # Crear un nuevo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avance por Etapas"

    # Encabezados
    headers = ["Etapa", "Presupuesto (COP)", "Gasto Ejecutado (COP)", "% Ejecutado", "Estado"]
    ws.append(headers)

    # Datos
    for etapa in etapas_con_avance:
        ws.append([
            etapa["nombre"],
            float(etapa["presupuesto"]),
            float(etapa["gasto"]),
            round(etapa["porcentaje"], 2),
            etapa["estado"]
        ])

    # Configurar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = f'attachment; filename=avance_etapas_proyecto_{project.id}.xlsx'
    wb.save(response)

    return response


@login_required
def project_update(request, project_id):
    """
    Vista para actualizar un proyecto
    PostgreSQL: UPDATE projects_project SET ... WHERE id = [project_id]
    """
    # Obtener proyecto específico del usuario actual
    project = get_object_or_404(Project, id=project_id, creado_por=request.user)

    workers = Worker.objects.all()
    if request.method == "POST":
        # Procesar formulario with datos existentes
        form = ProjectForm(request.POST, request.FILES, instance=project)
        selected_workers = request.POST.getlist("workers")
        if form.is_valid():
            try:
                # Actualizar en PostgreSQL
                form.save()
                # Calcular campos heredados automáticamente
                project.calculate_legacy_fields()
                # Calcular presupuesto actualizado
                project.presupuesto = project.calculate_final_budget()
                # Guardar con los campos calculados
                project.save()
                # Actualizar trabajadores asignados
                if selected_workers:
                    project.workers.set(selected_workers)
                else:
                    project.workers.clear()
                messages.success(request, "Proyecto actualizado exitosamente!")
                return redirect("projects:project_detail", project_id=project.id)
            except Exception as e:
                messages.error(request, f"❌ Error al actualizar el proyecto: {str(e)}")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        # Mostrar formulario con datos existentes
        form = ProjectForm(instance=project)

    return render(
        request,
        "projects/project_form.html",
        {
            "form": form,
            "project": project,
            "is_update": True,
            "workers": workers,
            "no_workers": not workers.exists(),
        },
    )


@project_owner_or_jefe_required
def project_duplicate(request, project_id):
    """
    Vista para duplicar un proyecto existente
    """
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == "POST":
        try:
            from .utils import duplicate_project
            new_project = duplicate_project(project)
            messages.success(
                request,
                f'✅ Proyecto duplicado exitosamente como "{new_project.name}"'
            )
            return redirect('projects:project_detail', project_id=new_project.id)
        except Exception as e:
            messages.error(request, f'❌ Error al duplicar el proyecto: {str(e)}')
            return redirect('projects:project_detail', project_id=project_id)
    
    return render(request, 'projects/project_duplicate_confirm.html', {
        'project': project
    })

@project_owner_or_jefe_required
def project_delete(request, project_id):
    """
    Vista para eliminar un proyecto
    - CONSTRUCTOR: Solo puede eliminar SUS proyectos
    - JEFE: Puede eliminar cualquier proyecto
    - COMERCIAL: Sin acceso
    """
    # Obtener proyecto (el decorador ya verificó los permisos)
    project = get_object_or_404(Project, id=project_id)

    if request.method == "POST":
        project.delete()
        messages.success(request, "Proyecto eliminado exitosamente!")
        return redirect("projects:project_list")

    return render(request, "projects/project_confirm_delete.html", {"project": project})


@login_required
def update_project_status(request, project_id):
    """
    Vista para actualizar el estado del proyecto (AJAX y POST)
    PostgreSQL: UPDATE projects_project SET estado = [new_status] WHERE id = [project_id]
    """
    if request.method == "POST":
        try:
            # Obtener proyecto - los jefes pueden modificar cualquier proyecto
            if request.user.role == User.JEFE or request.user.is_superuser:
                project = get_object_or_404(Project, id=project_id)
            else:
                # Los constructores solo pueden modificar sus propios proyectos
                project = get_object_or_404(Project, id=project_id, creado_por=request.user)

            # Verificar si es AJAX o formulario normal
            if request.content_type == "application/json":
                # Petición AJAX
                data = json.loads(request.body)
                new_status = data.get("status")
            else:
                # Formulario normal
                new_status = request.POST.get("status")

            # Validar estado
            valid_states = ["futuro", "en_proceso", "terminado"]
            if new_status not in valid_states:
                if request.content_type == "application/json":
                    return JsonResponse({"success": False, "error": "Estado inválido"})
                else:
                    messages.error(request, "❌ Estado inválido")
                    return redirect("projects:project_detail", project_id=project.id)

            # Actualizar estado
            project.estado = new_status
            project.save()

            # Responder según el tipo de petición
            if request.content_type == "application/json":
                return JsonResponse({"success": True, "status": new_status})
            else:
                messages.success(
                    request,
                    f'✅ Estado cambiado a "{project.get_estado_display()}" exitosamente',
                )
                return redirect("projects:project_detail", project_id=project.id)

        except Http404:
            # Proyecto no encontrado o usuario sin permisos
            if request.user.role == User.JEFE or request.user.is_superuser:
                error_msg = "El proyecto no existe."
            else:
                error_msg = "No tienes permisos para modificar este proyecto o el proyecto no existe."
            if request.content_type == "application/json":
                return JsonResponse({"success": False, "error": error_msg})
            else:
                messages.error(request, f"❌ {error_msg}")
                return redirect("projects:project_list")
                
        except Exception as e:
            if request.content_type == "application/json":
                return JsonResponse({"success": False, "error": str(e)})
            else:
                messages.error(request, f"❌ Error al cambiar estado: {str(e)}")
                return redirect("projects:project_list")

    return JsonResponse({"success": False, "error": "Método no permitido"})


@login_required
def recalculate_legacy_fields(request, project_id):
    """
    Vista para recalcular campos heredados de un proyecto específico via AJAX
    """
    if request.method == "POST":
        try:
            # Obtener proyecto
            project = get_object_or_404(Project, id=project_id, creado_por=request.user)

            # Calcular campos heredados
            project.calculate_legacy_fields()

            # Calcular presupuesto actualizado usando el método correcto
            project.presupuesto = project.calculate_final_budget()

            # Guardar cambios
            project.save()

            return JsonResponse(
                {
                    "success": True,
                    "message": "Campos recalculados exitosamente",
                    "data": {
                        "walls_area": float(project.walls_area),
                        "windows_area": float(project.windows_area),
                        "doors_count": project.doors_count,
                        "presupuesto": float(project.presupuesto),
                    },
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})


@login_required
def project_view(request):
    """
    Vista para mostrar proyectos con diseño de cards responsive
    """
    # Obtener el término de búsqueda
    search_query = request.GET.get("search", "")

    # Filtrar proyectos del usuario actual
    projects = Project.objects.filter(creado_por=request.user)

    # Aplicar búsqueda si se proporciona un término
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query)  # Buscar en nombre
            | Q(description__icontains=search_query)  # Buscar en descripción
            | Q(location_address__icontains=search_query)  # Buscar en dirección
        )

    context = {
        "projects": projects,
        "search_query": search_query,
    }

    return render(request, "projects/project_view.html", context)


@login_required
def worker_create(request):
    """
    Vista para crear un nuevo trabajador
    PostgreSQL: INSERT INTO projects_worker (...) VALUES (...)
    """
    if request.method == "POST":
        form = WorkerForm(request.POST)
        if form.is_valid():
            try:
                # Crear trabajador pero no guardar aún
                worker = form.save(commit=False)

                # Guardar en PostgreSQL
                worker.save()

                messages.success(
                    request, f'✅ Trabajador "{worker.name}" creado exitosamente!'
                )
                return redirect("projects:worker_list")
            except Exception as e:
                messages.error(request, f"❌ Error al crear el trabajador: {str(e)}")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        form = WorkerForm()
    return render(request, "projects/worker_form.html", {"form": form})


@login_required
def role_create(request):
    """
    Vista para crear un nuevo rol
    PostgreSQL: INSERT INTO projects_role (...) VALUES (...)
    """
    if request.method == "POST":
        form = RoleForm(request.POST)
        if form.is_valid():
            try:
                # Crear rol pero no guardar aún
                role = form.save(commit=False)

                # Guardar en PostgreSQL
                role.save()

                messages.success(request, f'✅ Rol "{role.name}" creado exitosamente!')
                return redirect("projects:role_list")
            except Exception as e:
                messages.error(request, f"❌ Error al crear el rol: {str(e)}")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        form = RoleForm()
    return render(request, "projects/role_form.html", {"form": form})


@login_required
def role_update(request, role_id):
    """
    Vista para actualizar un rol
    PostgreSQL: UPDATE projects_role SET ... WHERE id = [role_id]
    """
    role = get_object_or_404(Role, id=role_id)
    if request.method == "POST":
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            messages.success(request, f'Rol "{role.name}" actualizado exitosamente!')
            return redirect("projects:role_list")
    else:
        form = RoleForm(instance=role)
    return render(
        request,
        "projects/role_form.html",
        {"form": form, "role": role, "is_update": True},
    )


@login_required
def worker_list(request):
    """
    Vista para listar todos los trabajadores
    PostgreSQL: SELECT * FROM projects_worker
    """
    workers = Worker.objects.all().select_related('role')
    
    # Si se solicita exportar a Excel
    if 'export' in request.GET:
        # Crear un nuevo libro de trabajo de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Trabajadores"

        # Definir estilos
        header_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        # Encabezados
        headers = ['Nombre', 'Teléfono', 'Cédula', 'Dirección', 'Rol', 'EPS', 'ARL', 'Tipo de Sangre', 'Acudiente', 'Teléfono Acudiente']
        ws.append(headers)

        # Aplicar estilos a los encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Datos de trabajadores
        for worker in workers:
            ws.append([
                worker.name,
                worker.phone,
                worker.cedula,
                worker.direccion,
                worker.role.name if worker.role else '—',
                worker.eps,
                worker.arl,
                worker.blood_type or 'No especificado',
                worker.emergency_contact_name or 'No especificado',
                worker.emergency_contact_phone or 'No especificado'
            ])

        # Aplicar bordes y alineación a todas las celdas
        data_alignment = Alignment(horizontal='left', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = data_alignment

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Crear un buffer de bytes para guardar el archivo
        excel_file = io.BytesIO()
        
        # Guardar el libro de trabajo en el buffer
        wb.save(excel_file)
        
        # Preparar el buffer para lectura
        excel_file.seek(0)
        
        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="lista_trabajadores.xlsx"'
        return response

    return render(request, "projects/worker_list.html", {"workers": workers})

@login_required
def worker_edit(request, worker_id):
    """
    Vista para editar un trabajador existente
    """
    worker = get_object_or_404(Worker, id=worker_id)
    
    if request.method == "POST":
        form = WorkerForm(request.POST, instance=worker)
        if form.is_valid():
            worker = form.save()
            messages.success(request, f'Trabajador "{worker.name}" actualizado exitosamente.')
            return redirect('projects:worker_list')
    else:
        form = WorkerForm(instance=worker)
    
    return render(request, 'projects/worker_form.html', {
        'form': form,
        'worker': worker,
        'is_edit': True
    })


@login_required
def role_list(request):
    """
    Vista para listar todos los roles
    PostgreSQL: SELECT * FROM projects_role
    """
    roles = Role.objects.all()
    return render(request, "projects/role_list.html", {"roles": roles})


@login_required
def worker_delete(request, worker_id):
    """
    Vista para eliminar un trabajador
    PostgreSQL: DELETE FROM projects_worker WHERE id = [worker_id]
    """
    worker = get_object_or_404(Worker, id=worker_id)
    if request.method == "POST":
        worker.delete()
        messages.success(request, f'Trabajador "{worker.name}" eliminado exitosamente!')
        return redirect('projects:worker_list')
    return render(request, 'projects/worker_confirm_delete.html', {'worker': worker})

@login_required
def role_delete(request, role_id):
    """
    Vista para eliminar un rol
    PostgreSQL: DELETE FROM projects_role WHERE id = [role_id]
    """
    role = get_object_or_404(Role, id=role_id)
    if request.method == 'POST':
        role.delete()
        messages.success(request, f'Rol "{role.name}" eliminado exitosamente!')
        return redirect('projects:role_list')
    return render(request, 'projects/role_confirm_delete.html', {'role': role})

# Función para editar una entrada de material existente
@login_required
def editar_entrada_material(request, entrada_id):
    entrada = get_object_or_404(EntradaMaterial, id=entrada_id, proyecto__creado_por=request.user)
    
    if request.method == "POST":
        form = EntradaMaterialForm(request.POST, instance=entrada)
        if form.is_valid():
            form.save()
            messages.success(request, f'Entrada de {entrada.material.nombre} actualizada correctamente.')
            return redirect("projects:project_board", project_id=entrada.proyecto.id)
    else:
        form = EntradaMaterialForm(instance=entrada)
    
    return render(request, "projects/editar_entrada_material.html", {"form": form, "entrada": entrada})

@login_required
def borrar_entrada_material(request, entrada_id):
    """
    Borra una entrada de material y muestra un mensaje de éxito.
    """
    entrada = get_object_or_404(EntradaMaterial, id=entrada_id)

    if request.method == "POST":
        material_name = entrada.material.name  # usa el campo correcto
        entrada.delete()
        messages.success(
            request, f'Entrada de {material_name} eliminada correctamente.'
        )
        return redirect('projects:project_board', project_id=entrada.proyecto.id)
    
    # Redirige al tablero si se accede con GET
    return redirect('projects:project_board', project_id=entrada.proyecto.id)


# ===== VISTAS PARA CONSUMO DIARIO DE MATERIALES (RF17A) =====

@project_owner_or_jefe_required
def registrar_consumo_material(request, project_id):
    """
    Vista para registrar el consumo diario de materiales (RF17A)
    Con validación de stock insuficiente (RF17D)
    Se accede desde el calendario al seleccionar una fecha
    """
    project = get_object_or_404(Project, id=project_id)

    # Obtener fecha seleccionada del parámetro GET o usar hoy
    from django.utils import timezone
    fecha_seleccionada = request.GET.get('fecha', timezone.now().date())

    # Variables para manejar el error de stock insuficiente
    stock_insuficiente = False
    stock_disponible = None
    material_info = None

    if request.method == 'POST':
        form = ConsumoMaterialForm(request.POST, proyecto=project)

        if form.is_valid():
            try:
                consumo = form.save(commit=False)
                consumo.proyecto = project
                consumo.registrado_por = request.user
                consumo.save()

                messages.success(
                    request,
                    f'✅ Consumo registrado correctamente: {consumo.cantidad_consumida} {consumo.material.unit.symbol} '
                    f'de {consumo.material.name} para {consumo.componente_actividad}'
                )
                return redirect('projects:project_board', project_id=project.id)

            except Exception as e:
                messages.error(request, f'❌ Error al registrar consumo: {str(e)}')
        else:
            # Verificar si el error es de stock insuficiente (RF17D)
            if '__all__' in form.errors:
                error_msg = str(form.errors['__all__'][0])
                if 'Stock insuficiente' in error_msg:
                    stock_insuficiente = True
                    # Obtener información del stock desde el formulario
                    if hasattr(form, 'stock_disponible'):
                        stock_disponible = form.stock_disponible
                        material_info = {
                            'nombre': form.material_nombre,
                            'unidad': form.material_unidad,
                        }

                    messages.warning(request, f'⚠️ {error_msg}')
                else:
                    messages.error(request, error_msg)
            else:
                # Mostrar otros errores del formulario
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, f'{error}')
                        else:
                            messages.error(request, f'{field}: {error}')
    else:
        # Inicializar formulario with fecha seleccionada
        initial_data = {'fecha_consumo': fecha_seleccionada}
        form = ConsumoMaterialForm(initial=initial_data, proyecto=project)

    context = {
        'form': form,
        'project': project,
        'fecha_seleccionada': fecha_seleccionada,
        'stock_insuficiente': stock_insuficiente,
        'stock_disponible': stock_disponible,
        'material_info': material_info,
        'add_purchases_url': reverse("projects:registrar_entrada_material", kwargs={"project_id": project.id}),
    }

    return render(request, 'projects/registrar_consumo_material.html', context)


@project_owner_or_jefe_required
def listar_consumos_proyecto(request, project_id):
    """
    Vista para listar todos los consumos de materiales de un proyecto
    Con filtro por etapa del presupuesto (RF17B - Criterio 4)
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Obtener todos los consumos del proyecto
    consumos = ConsumoMaterial.objects.filter(
        proyecto=project
    ).select_related(
        'material__unit',
        'etapa_presupuesto',
        'registrado_por'
    ).order_by('-fecha_consumo', '-fecha_registro')
    
    # ✅ FILTRO POR ETAPA DEL PRESUPUESTO (RF17B - Criterio 4)
    etapa_filtro = request.GET.get('etapa')
    if etapa_filtro:
        try:
            etapa_id = int(etapa_filtro)
            consumos = consumos.filter(etapa_presupuesto_id=etapa_id)
        except (ValueError, TypeError):
            pass
    
    # Obtener todas las etapas para el filtro
    etapas_disponibles = BudgetSection.objects.all().order_by('order')
    
    # Calcular totales por material
    from django.db.models import Sum
    totales_por_material = consumos.values(
        'material__name',
        'material__unit__symbol'
    ).annotate(
        total_consumido=Sum('cantidad_consumida')
    ).order_by('material__name')

    context = {
        'project': project,
        'consumos': consumos,
        'etapas_disponibles': etapas_disponibles,
        'etapa_filtro': etapa_filtro,
        'totales_por_material': totales_por_material,
    }

    return render(request, 'projects/listar_consumos.html', context)

@project_owner_or_jefe_required
def obtener_consumos_fecha(request, project_id):
    """
    API endpoint para obtener consumos de una fecha específica (para el calendario)
    Retorna JSON con los consumos de la fecha
    """
    project = get_object_or_404(Project, id=project_id)
    fecha = request.GET.get('fecha')

    if not fecha:
        return JsonResponse({'error': 'Fecha no proporcionada'}, status=400)

    from .models import ConsumoMaterial
    consumos = ConsumoMaterial.objects.filter(
        proyecto=project,
        fecha_consumo=fecha
    ).select_related('material', 'material__unit').values(
        'id',
        'material__name',
        'material__sku',
        'cantidad_consumida',
        'material__unit__symbol',
        'componente_actividad',
        'responsable',
        'observaciones'
    )

    return JsonResponse({
        'fecha': fecha,
        'consumos': list(consumos),
        'total': consumos.count()
    })


@project_owner_or_jefe_required
def obtener_consumos_mes(request, project_id):
    """
    API endpoint para obtener todos los consumos de un mes específico (RF17C)
    Retorna JSON con los consumos agrupados por fecha para el calendario
    """
    project = get_object_or_404(Project, id=project_id)
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    if not mes or not anio:
        return JsonResponse({'error': 'Mes y año requeridos'}, status=400)

    try:
        mes = int(mes)
        anio = int(anio)
    except ValueError:
        return JsonResponse({'error': 'Mes y año deben ser números'}, status=400)

    from .models import ConsumoMaterial
    from datetime import date
    from collections import defaultdict

    # Obtener primer y último día del mes
    primer_dia = date(anio, mes, 1)
    if mes == 12:
        ultimo_dia = date(anio + 1, 1, 1)
    else:
        ultimo_dia = date(anio, mes + 1, 1)

    # Consultar consumos del mes
    consumos = ConsumoMaterial.objects.filter(
        proyecto=project,
        fecha_consumo__gte=primer_dia,
        fecha_consumo__lt=ultimo_dia
    ).select_related('material', 'material__unit').order_by('fecha_consumo')

    # Agrupar por fecha
    consumos_por_fecha = defaultdict(list)
    for consumo in consumos:
        fecha_str = consumo.fecha_consumo.strftime('%Y-%m-%d')
        consumos_por_fecha[fecha_str].append({
            'id': consumo.id,
            'material': consumo.material.name,
            'cantidad': float(consumo.cantidad_consumida),
            'unidad': consumo.material.unit.symbol,
            'actividad': consumo.componente_actividad,
            'responsable': consumo.responsable
        })

    return JsonResponse({
        'mes': mes,
        'anio': anio,
        'consumos_por_fecha': dict(consumos_por_fecha),
        'total_dias_con_registro': len(consumos_por_fecha)
    })


@project_owner_or_jefe_required
def editar_consumo_material(request, consumo_id):
    """
    Vista para editar un consumo de material existente
    Permite cambiar la etapa del presupuesto (RF17B - Criterio 5)
    """
    consumo = get_object_or_404(
        ConsumoMaterial.objects.select_related('proyecto', 'material__unit'),
        id=consumo_id
    )
    project = consumo.proyecto

    # Variables para manejar el error de stock insuficiente
    stock_insuficiente = False
    stock_disponible = None
    material_info = None

    if request.method == 'POST':
        form = ConsumoMaterialForm(request.POST, instance=consumo, proyecto=project)

        if form.is_valid():
            try:
                consumo_actualizado = form.save()

                messages.success(
                    request,
                    f'✅ Consumo actualizado correctamente: {consumo_actualizado.cantidad_consumida} '
                    f'{consumo_actualizado.material.unit.symbol} de {consumo_actualizado.material.name}'
                )
                return redirect('projects:listar_consumos_proyecto', project_id=project.id)

            except Exception as e:
                messages.error(request, f'❌ Error al actualizar consumo: {str(e)}')
        else:
            # Verificar si el error es de stock insuficiente
            if '__all__' in form.errors:
                error_msg = str(form.errors['__all__'][0])
                if 'Stock insuficiente' in error_msg:
                    stock_insuficiente = True
                    if hasattr(form, 'stock_disponible'):
                        stock_disponible = form.stock_disponible
                        material_info = {
                            'nombre': form.material_nombre,
                            'unidad': form.material_unidad,
                        }
                    messages.warning(request, f'⚠️ {error_msg}')
                else:
                    messages.error(request, error_msg)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, f'{error}')
                        else:
                            messages.error(request, f'{field}: {error}')
    else:
        form = ConsumoMaterialForm(instance=consumo, proyecto=project)

    context = {
        'form': form,
        'project': project,
        'consumo': consumo,
        'is_edit': True,
        'stock_insuficiente': stock_insuficiente,
        'stock_disponible': stock_disponible,
        'material_info': material_info,
        'add_purchases_url': reverse("projects:registrar_entrada_material", kwargs={"project_id": project.id}),
    }

    return render(request, 'projects/registrar_consumo_material.html', context)

@login_required
def eliminar_consumo_material(request, consumo_id):
    """
    Vista para eliminar un consumo de material
    """
    from .models import ConsumoMaterial
    consumo = get_object_or_404(ConsumoMaterial, id=consumo_id)
    project = consumo.proyecto
    project_id = project.id
    
    # Verificar permisos: JEFE o creador del proyecto
    if request.user.role != 'JEFE' and not request.user.is_superuser:
        if project.creado_por != request.user:
            raise PermissionDenied("No tienes permisos para eliminar consumos de este proyecto")

    if request.method == 'POST':
        try:
            material_name = consumo.material.name
            cantidad = consumo.cantidad_consumida
            unidad = consumo.material.unit.symbol

            # Al eliminar, el stock se restaura automáticamente en el modelo
            consumo.delete()

            messages.success(
                request,
                f'✅ Consumo eliminado correctamente: {cantidad} {unidad} de {material_name}. El stock ha sido restaurado.'
            )
        except Exception as e:
            messages.error(
                request,
                f'❌ Error al eliminar el consumo: {str(e)}'
            )

        # Redirigir al tablero del proyecto
        return redirect('projects:project_board', project_id=project_id)

    # Si no es POST, también redirigir al tablero
    messages.warning(request, 'Método no permitido')
    return redirect('projects:project_board', project_id=project_id)


# VISTAS PARA PRESUPUESTO DETALLADO

@role_required(User.CONSTRUCTOR, User.JEFE)
@login_required
def detailed_project_create(request):
    """
    Crea proyectos con presupuesto detallado completo.
    Usa las 23 secciones predefinidas (plantillas globales) sin modificarlas.
    """
    # Usar solo las secciones plantilla (sin proyecto asignado)
    sections = BudgetSection.objects.filter(project__isnull=True).order_by('order')
    workers = Worker.objects.all()

    if request.method == "POST":
        project_form = DetailedProjectForm(request.POST, request.FILES)
        selected_workers = request.POST.getlist("workers")

        if project_form.is_valid():
            try:
                from decimal import Decimal

                project = project_form.save(commit=False)
                project.creado_por = request.user
                project.built_area = Decimal("0")
                project.exterior_area = Decimal("0")
                project.columns_count = 0
                project.walls_area = Decimal("0")
                project.windows_area = Decimal("0")
                project.doors_count = 0
                project.doors_height = Decimal("2.1")
                
                # Usar el porcentaje de administración del formulario, o el de la sección plantilla como fallback
                admin_percentage = request.POST.get('administration_percentage')
                if admin_percentage:
                    try:
                        project.administration_percentage = Decimal(str(admin_percentage))
                    except (ValueError, TypeError):
                        # Si hay error, usar el valor de la sección plantilla
                        admin_section = sections.filter(order=21).first()
                        if admin_section and admin_section.percentage_value:
                            project.administration_percentage = admin_section.percentage_value
                else:
                    # Si no se proporciona, usar el valor de la sección plantilla
                    admin_section = sections.filter(order=21).first()
                    if admin_section and admin_section.percentage_value:
                        project.administration_percentage = admin_section.percentage_value
                
                project.save()

                # from .utils import create_default_budget_sections
                # create_default_budget_sections(project)
                
                if selected_workers:
                    project.workers.set(selected_workers)

                # Procesar cada sección (las plantillas base)
                items_processed = 0
                for section in sections:
                    section_form = BudgetSectionForm(section, project, request.POST)
                    if section_form.is_valid():
                        section_form.save(project)
                        items_processed += 1

                # Calcular presupuesto final
                presupuesto_calculado = project.calculate_final_budget()
                project.presupuesto = presupuesto_calculado
                project.save()

                messages.success(
                    request,
                    f'✅ Proyecto "{project.name}" creado exitosamente con presupuesto detallado. '
                    f'Presupuesto total: ${project.presupuesto:,.0f}',
                )

                return redirect("projects:detailed_budget_view", project_id=project.id)

            except Exception as e:
                messages.error(request, f"❌ Error al crear el proyecto: {str(e)}")

        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario.")
    else:
        project_form = DetailedProjectForm()
        # Inicializar el porcentaje de administración con el valor de la sección plantilla
        admin_section = sections.filter(order=21).first()
        if admin_section and admin_section.percentage_value:
            project_form.fields['administration_percentage'].initial = admin_section.percentage_value

    # Generar formularios por sección (render inicial)
    section_forms = [
        {"section": s, "form": BudgetSectionForm(s, None)} for s in sections
    ]

    return render(
        request,
        "projects/detailed_project_form.html",
        {
            "project_form": project_form,
            "section_forms": section_forms,
            "workers": workers,
            "no_workers": not workers.exists(),
        },
    )



@role_required(User.CONSTRUCTOR, User.JEFE)
@login_required
def detailed_budget_edit(request, project_id):
    """
    NUEVA VISTA SIMPLE: Editar presupuesto detallado
    """
    print(f"🔍 DEBUG: ===== INICIANDO detailed_budget_edit =====")
    print(f"🔍 DEBUG: Proyecto ID: {project_id}")
    print(f"🔍 DEBUG: Usuario: {request.user}")
    print(f"🔍 DEBUG: Rol usuario: {request.user.role}")
    
    project = get_object_or_404(Project, id=project_id)
    print(f"🔍 DEBUG: Proyecto encontrado: {project.name}")
    
    # Verificar permisos
    if request.user.role == User.CONSTRUCTOR and project.creado_por != request.user:
        print(f"🔍 DEBUG: ❌ Permisos denegados")
        raise PermissionDenied("No tienes permisos para editar este proyecto")
    
    print(f"🔍 DEBUG: ✅ Permisos verificados")
    print(f"🔍 DEBUG: Nueva vista simple para proyecto {project.id}")
    
    if request.method == "POST":
        print("🔍 DEBUG: Procesando formulario POST")
        
        # Procesar solo las cantidades que vienen en el POST
        items_updated = 0
        for key, value in request.POST.items():
            if key.startswith('quantity_'):
                item_id = key.replace('quantity_', '')
                try:
                    from decimal import Decimal
                    quantity = Decimal(str(value)) if value else Decimal('0')
                    
                    # Obtener el BudgetItem para usar su precio unitario
                    try:
                        budget_item = BudgetItem.objects.get(id=item_id)
                        
                        # Actualizar o crear ProjectBudgetItem (incluso con cantidad 0)
                        project_item, created = ProjectBudgetItem.objects.get_or_create(
                            project=project,
                            budget_item=budget_item,
                            defaults={
                                'quantity': quantity,
                                'unit_price': budget_item.unit_price
                            }
                        )
                        if not created:
                            project_item.quantity = quantity
                            project_item.unit_price = budget_item.unit_price
                            project_item.save()
                        items_updated += 1
                        print(f"✅ Actualizado ítem {item_id}: cantidad {quantity}, precio {budget_item.unit_price}")
                    except BudgetItem.DoesNotExist:
                        print(f"❌ BudgetItem {item_id} no existe")
                        continue
                except (ValueError, ProjectBudgetItem.DoesNotExist):
                    continue
        
        # Recalcular presupuesto
        presupuesto_calculado = project.calculate_final_budget()
        project.presupuesto = presupuesto_calculado
        project.save()
        
        print(f"🔍 DEBUG: {items_updated} ítems actualizados")
        print(f"🔍 DEBUG: Presupuesto recalculado: ${presupuesto_calculado:,.0f}")
        
        messages.success(request, f'✅ Presupuesto actualizado! {items_updated} ítems modificados.')
        return redirect("projects:detailed_budget_view", project_id=project.id)
    
    # Obtener TODAS las secciones (las 23) - optimizar consulta
    all_sections = BudgetSection.objects.select_related().prefetch_related('items').order_by('order')
    
    # Obtener ítems configurados del proyecto
    project_items = ProjectBudgetItem.objects.filter(project=project).select_related(
        'budget_item__section'
    ).order_by('budget_item__section__order', 'budget_item__order')
    
    # Crear diccionario de ítems por ID para búsqueda rápida
    project_items_dict = {item.budget_item.id: item for item in project_items}
    
    # Preparar datos para TODAS las secciones - optimizar consultas
    # Pre-cargar todos los ítems de una vez
    all_budget_items = BudgetItem.objects.filter(
        section__in=all_sections, 
        is_active=True
    ).select_related('section').order_by('section__order', 'order')
    
    # Crear diccionario de ítems por sección
    items_by_section = {}
    for item in all_budget_items:
        section_id = item.section.id
        if section_id not in items_by_section:
            items_by_section[section_id] = []
        items_by_section[section_id].append(item)
    
    # Importar Decimal antes de usarlo
    from decimal import Decimal
    
    sections_data = {}
    for section in all_sections:
        # Obtener ítems de esta sección del diccionario pre-cargado
        section_items = items_by_section.get(section.id, [])
        
        # Preparar ítems para esta sección
        items_for_section = []
        section_total = Decimal('0')  # Calcular total de la sección
        
        for budget_item in section_items:
            # Verificar si este ítem está configurado en el proyecto
            project_item = project_items_dict.get(budget_item.id)
            
            if project_item:
                # Ya está configurado, usar valores del proyecto
                # Convertir Decimal a float y luego a string sin formato para evitar problemas de renderizado
                quantity_value = float(project_item.quantity)
                # Convertir precio unitario a float y luego a string sin formato
                unit_price_value = float(project_item.unit_price)
                item_total = float(project_item.total_price)
                section_total += Decimal(str(item_total))
                
                items_for_section.append({
                    'budget_item': budget_item,
                    'project_item': project_item,
                    'quantity': quantity_value,
                    'quantity_str': str(quantity_value).rstrip('0').rstrip('.') if quantity_value % 1 == 0 else str(quantity_value),
                    'unit_price': unit_price_value,
                    'unit_price_str': str(unit_price_value).rstrip('0').rstrip('.') if unit_price_value % 1 == 0 else str(unit_price_value),
                    'total_price': item_total,
                    'is_configured': True
                })
                print(f"🔍 DEBUG: Ítem configurado {budget_item.description[:30]}: cantidad={project_item.quantity}, cantidad_str={str(quantity_value).rstrip('0').rstrip('.') if quantity_value % 1 == 0 else str(quantity_value)}, precio={project_item.unit_price}")
            else:
                # No está configurado, usar valores por defecto
                unit_price_default = float(budget_item.unit_price)
                items_for_section.append({
                    'budget_item': budget_item,
                    'project_item': None,
                    'quantity': 0.0,
                    'unit_price': unit_price_default,
                    'unit_price_str': str(unit_price_default).rstrip('0').rstrip('.') if unit_price_default % 1 == 0 else str(unit_price_default),
                    'total_price': 0.0,
                    'is_configured': False
                })
                print(f"🔍 DEBUG: Ítem no configurado {budget_item.description[:30]}: precio={budget_item.unit_price}")
        
        sections_data[section.id] = {
            'section': section,
            'items': items_for_section,
            'section_total': float(section_total)  # Total calculado de la sección
        }
    
    print(f"🔍 DEBUG: {len(sections_data)} secciones totales (todas las 23)")
    print(f"🔍 DEBUG: {len(project_items)} ítems configurados en el proyecto")
    
    # Debug: Verificar datos antes de enviar al template
    print(f"🔍 DEBUG: Datos para el template:")
    for section_id, data in sections_data.items():
        print(f"  Sección {data['section'].order} ({data['section'].name}): {len(data['items'])} ítems")
        for item in data['items']:
            print(f"    - {item['budget_item'].description[:30]}: cantidad={item['quantity']}, precio={item['unit_price']}, total={item['total_price']}")
    
    # Calcular costo directo y administración para el contexto
    costo_directo_total = Decimal('0')
    for section_id, data in sections_data.items():
        if not data['section'].is_percentage:
            for item in data['items']:
                if item['project_item']:
                    costo_directo_total += Decimal(str(item['total_price']))
    
    admin_percentage = project.administration_percentage / Decimal('100')
    administracion_automatica = costo_directo_total * admin_percentage
    
    try:
        context = {
            'project': project,
            'sections_data': sections_data,
            'total_budget': project.calculate_final_budget(),
            'administration_percentage': project.administration_percentage,
            'costo_directo': costo_directo_total,
            'administracion_automatica': administracion_automatica
        }
        
        print(f"🔍 DEBUG: ===== RENDERIZANDO TEMPLATE =====")
        print(f"🔍 DEBUG: Secciones en context: {len(sections_data)}")
        print(f"🔍 DEBUG: Presupuesto total: ${project.calculate_final_budget():,.0f}")
        
        return render(request, "projects/detailed_budget_edit.html", context)
        
    except Exception as e:
        print(f"❌ ERROR en detailed_budget_edit: {str(e)}")
        print(f"❌ Tipo de error: {type(e).__name__}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        
        messages.error(request, f"❌ Error al cargar el presupuesto: {str(e)}")
        return redirect("projects:project_detail", project_id=project.id)


@role_required(User.CONSTRUCTOR, User.JEFE)
@login_required
def detailed_budget_view(request, project_id):
    """
    Vista para ver el presupuesto detallado de un proyecto
    OPTIMIZADA: Solo carga secciones con ítems configurados
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if request.user.role == User.CONSTRUCTOR and project.creado_por != request.user:
        raise PermissionDenied("No tienes permisos para ver este proyecto")
    
    # Obtener secciones con ítems configurados Y secciones porcentuales (como Administración)
    # Usar Q objects para combinar ambas condiciones en una sola consulta
    from django.db.models import Q
    sections = BudgetSection.objects.filter(
        Q(items__projectbudgetitem__project=project) | 
        Q(is_percentage=True, project__isnull=True)
    ).distinct().order_by('order')
    
    # Obtener todos los ProjectBudgetItem del proyecto de una vez
    project_items_dict = {
        item.budget_item_id: item 
        for item in ProjectBudgetItem.objects.filter(project=project).select_related('budget_item')
    }
    
    section_data = []
    total_budget = 0
    costo_directo_total = 0
    
    for section in sections:
        items = []
        section_total = 0
        
        # Si es una sección porcentual, no tiene ítems individuales
        if section.is_percentage:
            # Para secciones porcentuales, no hay ítems que mostrar
            # El cálculo se hace automáticamente en el resumen
            section_data.append({
                'section': section,
                'items': [],
                'total': 0  # Se calcula automáticamente en el resumen
            })
        else:
            # Solo obtener ítems que están configurados en el proyecto
            section_items = BudgetItem.objects.filter(
                section=section, 
                is_active=True,
                projectbudgetitem__project=project
            ).order_by('order')
            
            for item in section_items:
                project_item = project_items_dict.get(item.id)
                item_total = project_item.total_price if project_item else 0
                
                items.append({
                    'item': item,
                    'project_item': project_item,
                    'total': item_total
                })
                section_total += item_total
            
            section_data.append({
                'section': section,
                'items': items,
                'total': section_total
            })
            costo_directo_total += section_total
    
    # Calcular administración automática usando el porcentaje del proyecto
    from decimal import Decimal
    admin_percentage = project.administration_percentage / Decimal('100')
    administracion_automatica = costo_directo_total * admin_percentage
    
    # ✅ USAR EL CÁLCULO CORRECTO QUE INCLUYE ADMINISTRACIÓN
    total_budget = project.calculate_final_budget()
    
    # Agrupar secciones en macro-secciones lógicas
    macro_sections = {
        'Obra Negra': {
            'name': 'Obra Negra',
            'description': 'Cimientos, estructura, columnas, vigas',
            'icon': 'fas fa-building',
            'color': 'primary',
            'sections': [],
            'total': Decimal('0')
        },
        'Obra Blanca': {
            'name': 'Obra Blanca',
            'description': 'Mampostería, revoques, instalaciones básicas',
            'icon': 'fas fa-paint-roller',
            'color': 'info',
            'sections': [],
            'total': Decimal('0')
        },
        'Acabados': {
            'name': 'Acabados',
            'description': 'Pisos, pintura, carpintería, ventanas, puertas',
            'icon': 'fas fa-hammer',
            'color': 'success',
            'sections': [],
            'total': Decimal('0')
        },
        'Instalaciones': {
            'name': 'Instalaciones',
            'description': 'Hidráulicas, eléctricas, gas',
            'icon': 'fas fa-plug',
            'color': 'warning',
            'sections': [],
            'total': Decimal('0')
        },
        'Exteriores': {
            'name': 'Exteriores',
            'description': 'Fachada, zonas comunes, paisajismo',
            'icon': 'fas fa-tree',
            'color': 'secondary',
            'sections': [],
            'total': Decimal('0')
        },
        'Otros': {
            'name': 'Otros',
            'description': 'Preliminares, supervisión, administración, impuestos',
            'icon': 'fas fa-list',
            'color': 'dark',
            'sections': [],
            'total': Decimal('0')
        }
    }
    
    # Mapeo de secciones a macro-secciones según su order
    section_to_macro = {
        # Obra Negra
        2: 'Obra Negra',   # EXCAVACIONES Y LLENOS
        3: 'Obra Negra',   # CIMENTACIONES
        4: 'Obra Negra',   # ESTRUCTURA
        5: 'Obra Negra',   # ACERO
        
        # Obra Blanca
        6: 'Obra Blanca',  # MAMPOSTERÍA
        7: 'Obra Blanca',  # ESTUCOS, REVOQUES, PINTURAS, DRYWALL
        
        # Acabados
        8: 'Acabados',     # PISOS, ENCHAPES Y SÓCALOS
        9: 'Acabados',     # JUNTAS
        10: 'Acabados',    # CARPINTERÍA METÁLICA Y VIDRIOS
        11: 'Acabados',    # CARPINTERÍA EN MADERA, MESONES Y ACCESORIOS
        
        # Instalaciones
        12: 'Instalaciones',  # APARATOS SANITARIOS E INSTALACIONES HIDRÁULICAS
        13: 'Instalaciones',  # INSTALACIONES ELÉCTRICAS Y REDES DE GAS
        
        # Exteriores
        14: 'Exteriores',  # CUBIERTAS
        17: 'Exteriores',  # PISOS EXTERIORES
        19: 'Exteriores',  # PAISAJISMO
        
        # Otros (por defecto)
        1: 'Otros',        # PRELIMINARES Y MANTENIMIENTOS
        15: 'Otros',       # ELECTRODOMÉSTICOS
        16: 'Otros',       # IMPERMEABILIZACIONES
        18: 'Otros',       # LIMPIEZA GENERAL Y VIDRIOS
        20: 'Otros',       # SUPERVISIÓN Y CONTROL DE OBRA
        21: 'Otros',       # ADMINISTRACIÓN
        22: 'Otros',       # IMPUESTOS Y ESCRITURACIONES
        23: 'Otros',       # ESTUDIOS Y DISEÑO
    }
    
    # Agrupar secciones
    for section_item in section_data:
        section_order = section_item['section'].order
        macro_key = section_to_macro.get(section_order, 'Otros')
        
        # Calcular total de la sección
        if section_item['section'].is_percentage and section_item['section'].order == 21:
            section_total = administracion_automatica
        else:
            section_total = Decimal(str(section_item['total']))
        
        macro_sections[macro_key]['sections'].append(section_item)
        macro_sections[macro_key]['total'] += section_total
    
    # Convertir totales a float para el template
    for macro_key in macro_sections:
        macro_sections[macro_key]['total'] = float(macro_sections[macro_key]['total'])
    
    # Ordenar macro-secciones según el orden deseado
    macro_sections_ordered = [
        macro_sections['Obra Negra'],
        macro_sections['Obra Blanca'],
        macro_sections['Acabados'],
        macro_sections['Instalaciones'],
        macro_sections['Exteriores'],
        macro_sections['Otros']
    ]
    
    context = {
        'project': project,
        'section_data': section_data,
        'macro_sections': macro_sections_ordered,
        'total_budget': total_budget,
        'administration_percentage': project.administration_percentage,
        'costo_directo': costo_directo_total,
        'administracion_automatica': administracion_automatica
    }
    
    return render(request, "projects/detailed_budget_view.html", context)


@login_required
def update_project_image(request, project_id):
    """
    Vista para actualizar la imagen del proyecto
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if request.user.role != User.JEFE and not request.user.is_superuser and project.creado_por != request.user:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar este proyecto'})
    
    # Obtener la imagen del request
    imagen_proyecto = request.FILES.get('imagen_proyecto')
    if not imagen_proyecto:
        return JsonResponse({'success': False, 'error': 'No se proporcionó ninguna imagen'})
    
    try:
        # Actualizar la imagen del proyecto
        project.imagen_proyecto = imagen_proyecto
        project.save()
        
        # Verificar que la imagen se guardó correctamente
        if project.imagen_proyecto and hasattr(project.imagen_proyecto, 'url'):
            image_url = project.imagen_proyecto.url
        else:
            return JsonResponse({'success': False, 'error': 'Error al guardar la imagen'})
        
        return JsonResponse({
            'success': True,
            'image_url': image_url,
            'message': 'Imagen actualizada correctamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def update_administration_percentage(request, project_id):
    """
    Vista para actualizar el porcentaje de administración del proyecto
    Solo JEFE, superuser o el creador del proyecto pueden modificar
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos: JEFE, superuser o creador del proyecto
    if request.user.role != User.JEFE and not request.user.is_superuser and project.creado_por != request.user:
        return JsonResponse({'success': False, 'error': 'No tienes permisos para modificar este proyecto'})
    
    # Obtener el porcentaje del request
    try:
        percentage = float(request.POST.get('percentage', 0))
        
        # Validar que esté en el rango 0-100
        if percentage < 0 or percentage > 100:
            return JsonResponse({'success': False, 'error': 'El porcentaje debe estar entre 0 y 100'})
        
        # Actualizar el porcentaje del proyecto
        from decimal import Decimal
        project.administration_percentage = Decimal(str(percentage))
        project.save()
        
        # Recalcular el presupuesto total
        total_budget = project.calculate_final_budget()
        
        return JsonResponse({
            'success': True,
            'percentage': float(project.administration_percentage),
            'total_budget': float(total_budget),
            'message': f'Porcentaje de administración actualizado a {percentage}%'
        })
        
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Porcentaje inválido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@role_required(User.JEFE)
@login_required
def budget_management(request):
    """
    Vista para gestionar precios unitarios (solo JEFE)
    """
    sections = BudgetSection.objects.filter(project__isnull=True).order_by('order')
    section_data = []
    
    for section in sections:
        items = BudgetItem.objects.filter(section=section).order_by('order')
        section_data.append({
            'section': section,
            'items': items
        })
    
    context = {
        'section_data': section_data
    }
    
    return render(request, "projects/budget_management.html", context)


@role_required(User.JEFE)
@login_required
def update_section_percentage(request, section_id):
    """
    Vista para actualizar el porcentaje de una sección plantilla (solo JEFE)
    Solo se puede actualizar secciones plantilla (project=None) y solo la sección 21 (Administración)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    section = get_object_or_404(BudgetSection, id=section_id)
    
    # Solo permitir actualizar secciones plantilla (project=None)
    if section.project is not None:
        return JsonResponse({'success': False, 'error': 'Solo se pueden actualizar secciones plantilla'})
    
    # Solo permitir actualizar la sección 21 (Administración)
    if section.order != 21:
        return JsonResponse({'success': False, 'error': 'Solo se puede actualizar el porcentaje de la sección de Administración'})
    
    # Obtener el porcentaje del request
    try:
        percentage = float(request.POST.get('percentage', 0))
        
        # Validar que esté en el rango 0-100
        if percentage < 0 or percentage > 100:
            return JsonResponse({'success': False, 'error': 'El porcentaje debe estar entre 0 y 100'})
        
        # Actualizar el porcentaje de la sección
        from decimal import Decimal
        section.percentage_value = Decimal(str(percentage))
        section.save()
        
        return JsonResponse({
            'success': True,
            'percentage': float(section.percentage_value),
            'message': f'Porcentaje de administración actualizado a {percentage}%'
        })
        
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Porcentaje inválido'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@role_required(User.JEFE)
@login_required
def budget_item_update(request, item_id):
    """
    Vista para actualizar un ítem del presupuesto (solo JEFE)
    """
    item = get_object_or_404(BudgetItem, id=item_id)
    
    if request.method == "POST":
        form = BudgetManagementForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Ítem "{item.description[:50]}" actualizado exitosamente!')
            return redirect("projects:budget_management")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        form = BudgetManagementForm(instance=item)
    
    context = {
        'item': item,
        'form': form,
        'title': 'Actualizar Precio'
    }
    
    return render(request, "projects/budget_item_price_form.html", context)


# VISTAS PARA GESTIÓN DE ÍTEMS DEL PRESUPUESTO

@role_required(User.JEFE)
@login_required
def budget_items_list(request):
    """
    Vista para listar todos los ítems del presupuesto con opciones de gestión
    """
    sections = BudgetSection.objects.all().order_by('order')
    section_data = []
    
    for section in sections:
        items = BudgetItem.objects.filter(section=section).order_by('order')
        section_data.append({
            'section': section,
            'items': items
        })
    
    context = {
        'section_data': section_data
    }
    
    return render(request, "projects/budget_items_list.html", context)


@role_required(User.JEFE)
@login_required
def budget_item_create(request):
    """
    Vista para crear un nuevo ítem del presupuesto
    """
    if request.method == "POST":
        form = BudgetItemCreateForm(request.POST)
        if form.is_valid():
            # Asignar orden automáticamente (siempre al final de la sección)
            section = form.cleaned_data['section']
            last_order = BudgetItem.objects.filter(section=section).aggregate(
                max_order=Max('order')
            )['max_order'] or 0
            form.instance.order = last_order + 1
            
            form.save()
            messages.success(request, f'✅ Ítem "{form.instance.description[:50]}" creado exitosamente!')
            return redirect("projects:budget_items_list")
        else:
            messages.error(request, "❌ Por favor corrige los errores en el formulario")
    else:
        form = BudgetItemCreateForm()
    
    context = {
        'form': form,
        'title': 'Crear Nuevo Ítem'
    }
    
    return render(request, "projects/budget_item_form.html", context)


@role_required(User.JEFE)
@login_required
def budget_item_edit(request, item_id):
    """
    Vista para editar un ítem del presupuesto
    """
    item = get_object_or_404(BudgetItem, id=item_id)
    
    if request.method == "POST":
        form = BudgetItemEditForm(request.POST, instance=item)
        if form.is_valid():
            updated_item = form.save()
            print(f"🔍 DEBUG: Ítem actualizado - ID: {updated_item.id}")
            print(f"🔍 DEBUG: Precio anterior: {item.unit_price}")
            print(f"🔍 DEBUG: Precio nuevo: {updated_item.unit_price}")
            print(f"🔍 DEBUG: Descripción: {updated_item.description}")
            print(f"🔍 DEBUG: Activo: {updated_item.is_active}")
            messages.success(request, f'✅ Ítem "{updated_item.description[:50]}" actualizado exitosamente!')
            return redirect("projects:budget_items_list")
        else:
            # Debug: mostrar errores específicos
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            messages.error(request, f"❌ Errores en el formulario: {'; '.join(error_messages)}")
    else:
        form = BudgetItemEditForm(instance=item)
    
    context = {
        'form': form,
        'item': item,
        'title': 'Editar Ítem'
    }
    
    return render(request, "projects/budget_item_form.html", context)


@role_required(User.JEFE)
@login_required
def budget_item_delete(request, item_id):
    """
    Vista para eliminar un ítem del presupuesto
    """
    try:
        item = BudgetItem.objects.get(id=item_id)
    except BudgetItem.DoesNotExist:
        messages.error(request, f'❌ El ítem con ID {item_id} no existe o ya fue eliminado.')
        return redirect("projects:budget_items_list")
    
    if request.method == "POST":
        item_name = item.description[:50]
        item.delete()
        messages.success(request, f'✅ Ítem "{item_name}" eliminado exitosamente!')
        return redirect("projects:budget_items_list")
    
    context = {
        'item': item
    }
    
    return render(request, "projects/budget_item_delete_confirm.html", context)


@role_required(User.JEFE)
@login_required
def budget_item_toggle(request, item_id):
    """
    Vista para activar/desactivar un ítem del presupuesto
    """
    item = get_object_or_404(BudgetItem, id=item_id)
    
    if request.method == "POST":
        item.is_active = not item.is_active
        item.save()
        
        status = "activado" if item.is_active else "desactivado"
        messages.success(request, f'✅ Ítem "{item.description[:50]}" {status} exitosamente!')
    
    return redirect("projects:budget_items_list")


def calculate_detailed_budget_total(project):
    """
    FUNCIÓN ESPECÍFICA PARA FORMULARIO DETALLADO
    Suma ítems del presupuesto + administración automática (usando porcentaje del proyecto)
    """
    from decimal import Decimal
    
    costo_directo = Decimal('0')
    administracion_manual = Decimal('0')
    
    # Separar costo directo de administración manual
    project_items = ProjectBudgetItem.objects.filter(project=project)
    for item in project_items:
        # Si es de la sección 21 (Administración), contarlo por separado
        if item.budget_item.section.order == 21:
            administracion_manual += item.total_price
        else:
            costo_directo += item.total_price
    
    # Calcular administración automática usando el porcentaje del proyecto (default 12%)
    admin_percentage = project.administration_percentage / Decimal('100')
    administracion_automatica = costo_directo * admin_percentage
    
    # Total = Costo Directo + Administración Automática + Administración Manual
    total = costo_directo + administracion_automatica + administracion_manual
    
    return total


@role_required(User.JEFE, User.CONSTRUCTOR)
@login_required
def project_workers(request, project_id):
    """
    Vista para gestionar trabajadores asignados a un proyecto
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos
    if request.user.role == User.CONSTRUCTOR and project.creado_por != request.user:
        raise PermissionDenied("No tienes permisos para gestionar trabajadores de este proyecto")
    
    if request.method == "POST":
        # Obtener trabajadores seleccionados
        selected_workers = request.POST.getlist('workers')
        
        # Actualizar trabajadores del proyecto
        project.workers.set(selected_workers)
        
        messages.success(request, f'✅ Trabajadores actualizados exitosamente!')
        return redirect("projects:project_detail", project_id=project.id)
    
    # Obtener todos los trabajadores disponibles
    all_workers = Worker.objects.all().order_by('name')
    
    # Obtener trabajadores actualmente asignados
    assigned_workers = project.workers.all()
    
    context = {
        'project': project,
        'all_workers': all_workers,
        'assigned_workers': assigned_workers,
    }
    
    return render(request, "projects/project_workers.html", context)
@login_required
def project_graficos(request, project_id):
    """
    Vista principal para mostrar los gráficos comparativos de presupuesto vs gasto real.
    RF18: Control Diario de Gasto y Avance - Visualización de Gráficos
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar permisos (opcional)
    # if not request.user.has_perm('projects.view_project'):
    #     return redirect('projects:project_list')
    
    context = {
        'project': project,
    }
    
    return render(request, 'projects/graficos_proyecto.html', context)

# ===== VISTAS PARA RF18 - GRÁFICOS =====

@login_required
def project_graficos(request, project_id):
    """
    Vista principal para mostrar los gráficos comparativos de presupuesto vs gasto real.
    RF18: Control Diario de Gasto y Avance - Visualización de Gráficos
    """
    project = get_object_or_404(Project, id=project_id)
    
    context = {
        'project': project,
    }
    
    return render(request, 'projects/graficos_proyecto.html', context)


@login_required
def api_datos_graficos(request, project_id):
    """API que devuelve datos JSON para los gráficos"""
    try:
        project = get_object_or_404(Project, id=project_id)
        
        periodo = request.GET.get('periodo', 'todo')
        fecha_fin = datetime.now().date()
        
        if periodo == 'mes':
            fecha_inicio = fecha_fin - timedelta(days=30)
        elif periodo == 'trimestre':
            fecha_inicio = fecha_fin - timedelta(days=90)
        else:
            fecha_inicio = project.fecha_creacion.date() if project.fecha_creacion else (fecha_fin - timedelta(days=365))
        
        # Consolidado
        presupuesto_total = float(project.presupuesto or 0)
        gasto_total = float(project.presupuesto_gastado_calculado or 0)
        
        # Por material
        materiales_dict = {}
        
        # PRIMERO: Obtener todos los materiales del proyecto con sus entradas
        entradas = EntradaMaterial.objects.filter(proyecto=project).select_related('material')
        
        print(f"Total entradas en el proyecto: {entradas.count()}")
        
        for entrada in entradas:
            mid = entrada.material.id
            if mid not in materiales_dict:
                materiales_dict[mid] = {
                    'material': entrada.material.name,
                    'presupuesto': 0,
                    'gasto_real': 0,
                }
            # Presupuesto = todo lo que se ha comprado (sin importar fecha)
            precio = float(entrada.material.unit_cost or 0)
            cantidad = float(entrada.cantidad or 0)
            materiales_dict[mid]['presupuesto'] += precio * cantidad
        
        # SEGUNDO: Calcular consumos en el período seleccionado
        consumos = ConsumoMaterial.objects.filter(
            proyecto=project,
            fecha_consumo__gte=fecha_inicio,
            fecha_consumo__lte=fecha_fin
        ).select_related('material')
        
        print(f"Consumos en período: {consumos.count()}")
        
        for consumo in consumos:
            mid = consumo.material.id
            if mid not in materiales_dict:
                # Si no hay entrada, crear con presupuesto 0
                materiales_dict[mid] = {
                    'material': consumo.material.name,
                    'presupuesto': 0,
                    'gasto_real': 0,
                }
            precio = float(consumo.material.unit_cost or 0)
            cantidad = float(consumo.cantidad_consumida or 0)
            materiales_dict[mid]['gasto_real'] += precio * cantidad
        
        datos_por_material = list(materiales_dict.values())
        
        # Evolución temporal
        evolucion = []
        consumos_ord = ConsumoMaterial.objects.filter(
            proyecto=project,
            fecha_consumo__gte=fecha_inicio,
            fecha_consumo__lte=fecha_fin
        ).select_related('material').order_by('fecha_consumo')
        
        gasto_acum = 0
        fecha_act = None
        gasto_dia = 0
        
        for consumo in consumos_ord:
            if fecha_act and consumo.fecha_consumo != fecha_act:
                gasto_acum += gasto_dia
                evolucion.append({
                    'fecha': fecha_act.strftime('%Y-%m-%d'),
                    'gasto_dia': round(gasto_dia, 2),
                    'gasto_acumulado': round(gasto_acum, 2),
                })
                gasto_dia = 0
            
            precio = float(consumo.material.unit_cost or 0)
            cantidad = float(consumo.cantidad_consumida or 0)
            gasto_dia += precio * cantidad
            fecha_act = consumo.fecha_consumo
        
        if fecha_act:
            gasto_acum += gasto_dia
            evolucion.append({
                'fecha': fecha_act.strftime('%Y-%m-%d'),
                'gasto_dia': round(gasto_dia, 2),
                'gasto_acumulado': round(gasto_acum, 2),
            })
        
        response_data = {
            'consolidado': {
                'presupuesto': presupuesto_total,
                'gasto_real': gasto_total,
            },
            'por_material': datos_por_material,
            'evolucion_temporal': evolucion,
            'periodo': periodo,
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
        }
        
        print(f"Respuesta: {len(datos_por_material)} materiales, {len(evolucion)} días")
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'consolidado': {'presupuesto': 0, 'gasto_real': 0},
            'por_material': [],
            'evolucion_temporal': [],
        }, status=200)

@role_required(User.JEFE)
@login_required
def export_budget_to_excel(request, project_id):
    """
    Vista para exportar el presupuesto detallado a Excel
    Solo accesible para usuarios con rol JEFE
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Verificar que el proyecto tenga presupuesto detallado
    project_items_count = project.budget_items.count()
    print(f"🔍 DEBUG Excel Export - Total ProjectBudgetItems: {project_items_count}")
    
    if not project.budget_items.exists():
        print(f"❌ DEBUG Excel Export - No hay presupuesto detallado para proyecto {project.id}")
        messages.error(request, "❌ Este proyecto no tiene presupuesto detallado configurado.")
        return redirect("projects:project_board", project_id=project.id)
    
    print(f"✅ DEBUG Excel Export - Proyecto tiene presupuesto detallado, procediendo...")
    
    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    
    # Estilos para el formato profesional
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    subheader_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    normal_font = Font(name='Arial', size=10)
    currency_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', size=11, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # Eliminar la hoja por defecto
    workbook.remove(workbook.active)
    
    # Obtener todas las secciones del presupuesto (no solo las que tienen items configurados)
    all_sections = BudgetSection.objects.all().order_by('order')
    print(f"🔍 DEBUG Excel Export - Total secciones en sistema: {all_sections.count()}")
    
    # Filtrar solo las secciones que tienen ítems configurados en este proyecto
    sections_with_data = []
    for section in all_sections:
        items_in_section = ProjectBudgetItem.objects.filter(
            project=project,
            budget_item__section=section,
            quantity__gt=0  # Solo ítems con cantidad mayor a 0
        ).count()
        if items_in_section > 0:
            sections_with_data.append(section)
            print(f"  ✅ Sección {section.order}: {section.name} - {items_in_section} ítems")
        else:
            print(f"  ❌ Sección {section.order}: {section.name} - Sin ítems configurados")
    
    print(f"🔍 DEBUG Excel Export - Secciones con datos: {len(sections_with_data)}")
    
    # Obtener todos los ProjectBudgetItem del proyecto
    project_items = ProjectBudgetItem.objects.filter(project=project).select_related('budget_item', 'budget_item__section')
    project_items_dict = {item.budget_item_id: item for item in project_items}
    
    print(f"🔍 DEBUG Excel Export - ProjectBudgetItems encontrados: {project_items.count()}")
    for item in project_items:
        print(f"  - {item.budget_item.code or 'Sin código'}: {item.budget_item.description[:50]} - Cantidad: {item.quantity} - Precio: {item.unit_price}")
    
    # Variables para el resumen
    total_sections = {}
    grand_total = 0
    
    # Verificar si hay secciones con datos
    if not sections_with_data:
        print(f"❌ DEBUG Excel Export - No hay secciones con datos, pero el proyecto tiene {project_items.count()} items")
        messages.warning(request, "⚠️ Este proyecto tiene presupuesto configurado pero sin cantidades. Por favor configure las cantidades primero.")
        return redirect("projects:detailed_budget_edit", project_id=project.id)
    
    # Crear una hoja por cada sección
    for section in sections_with_data:
        print(f"🔍 DEBUG Excel Export - Procesando sección: {section.order}. {section.name}")
        
        # Crear hoja para la sección
        sheet_name = f"{section.order}. {section.name[:25]}"  # Limitar nombre de hoja
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Configurar encabezado del proyecto
        worksheet.merge_cells('A1:F1')
        worksheet['A1'] = f"PRESUPUESTO DETALLADO - {project.name.upper()}"
        worksheet['A1'].font = Font(name='Arial', size=14, bold=True)
        worksheet['A1'].alignment = center_alignment
        worksheet['A1'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        worksheet.merge_cells('A2:F2')
        worksheet['A2'] = f"Fecha de exportación: {get_colombia_time().strftime('%d/%m/%Y %H:%M')}"
        worksheet['A2'].font = Font(name='Arial', size=10, italic=True)
        worksheet['A2'].alignment = center_alignment
        
        # Encabezado de la sección
        row = 4
        worksheet.merge_cells(f'A{row}:F{row}')
        worksheet[f'A{row}'] = f"SECCIÓN {section.order}: {section.name.upper()}"
        worksheet[f'A{row}'].font = header_font
        worksheet[f'A{row}'].fill = header_fill
        worksheet[f'A{row}'].alignment = center_alignment
        
        # Encabezados de columnas
        row += 1
        headers = ['Código', 'Descripción', 'Unidad', 'Cantidad', 'Precio Unitario (COP)', 'Total (COP)']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Obtener ítems de la sección configurados en el proyecto
        section_items = BudgetItem.objects.filter(
            section=section, 
            is_active=True,
            projectbudgetitem__project=project
        ).order_by('order')
        
        print(f"🔍 DEBUG Excel Export - Ítems en sección {section.name}: {section_items.count()}")
        
        section_total = 0
        row += 1
        items_added = 0
        
        for item in section_items:
            project_item = project_items_dict.get(item.id)
            print(f"  🔍 DEBUG - Item {item.code or 'Sin código'}: {item.description[:30]}")
            print(f"    - ProjectItem encontrado: {project_item is not None}")
            if project_item:
                print(f"    - Cantidad: {project_item.quantity}, Precio: {project_item.unit_price}")
            
            if project_item and project_item.quantity > 0:
                print(f"    ✅ Agregando ítem al Excel")
                items_added += 1
                
                # Código
                worksheet.cell(row=row, column=1).value = item.code or f"{section.order}.{item.order}"
                worksheet.cell(row=row, column=1).font = normal_font
                worksheet.cell(row=row, column=1).border = border
                
                # Descripción
                worksheet.cell(row=row, column=2).value = item.description
                worksheet.cell(row=row, column=2).font = normal_font
                worksheet.cell(row=row, column=2).border = border
                
                # Unidad
                worksheet.cell(row=row, column=3).value = item.unit
                worksheet.cell(row=row, column=3).font = normal_font
                worksheet.cell(row=row, column=3).alignment = center_alignment
                worksheet.cell(row=row, column=3).border = border
                
                # Cantidad
                worksheet.cell(row=row, column=4).value = float(project_item.quantity)
                worksheet.cell(row=row, column=4).font = normal_font
                worksheet.cell(row=row, column=4).alignment = right_alignment
                worksheet.cell(row=row, column=4).border = border
                worksheet.cell(row=row, column=4).number_format = '#,##0.000'
                
                # Precio Unitario
                worksheet.cell(row=row, column=5).value = float(project_item.unit_price)
                worksheet.cell(row=row, column=5).font = currency_font
                worksheet.cell(row=row, column=5).alignment = right_alignment
                worksheet.cell(row=row, column=5).border = border
                worksheet.cell(row=row, column=5).number_format = '"$"#,##0'
                
                # Total
                item_total = float(project_item.total_price)
                worksheet.cell(row=row, column=6).value = item_total
                worksheet.cell(row=row, column=6).font = currency_font
                worksheet.cell(row=row, column=6).alignment = right_alignment
                worksheet.cell(row=row, column=6).border = border
                worksheet.cell(row=row, column=6).number_format = '"$"#,##0'
                
                section_total += item_total
                row += 1
        
        # Total de la sección
        row += 1
        worksheet.merge_cells(f'A{row}:E{row}')
        worksheet[f'A{row}'] = f"TOTAL SECCIÓN {section.order}: {section.name.upper()}"
        worksheet[f'A{row}'].font = total_font
        worksheet[f'A{row}'].alignment = right_alignment
        worksheet[f'A{row}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        worksheet[f'A{row}'].border = border
        
        worksheet.cell(row=row, column=6).value = section_total
        worksheet.cell(row=row, column=6).font = total_font
        worksheet.cell(row=row, column=6).alignment = right_alignment
        worksheet.cell(row=row, column=6).number_format = '"$"#,##0'
        worksheet.cell(row=row, column=6).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        worksheet.cell(row=row, column=6).border = border
        
        print(f"🔍 DEBUG Excel Export - Sección {section.name} completada:")
        print(f"  - Ítems agregados: {items_added}")
        print(f"  - Total de sección: ${section_total:,.0f}")
        
        # Guardar total para el resumen
        total_sections[section.name] = section_total
        grand_total += section_total
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 20
    
    # Crear hoja de RESUMEN
    summary_sheet = workbook.create_sheet(title="RESUMEN", index=0)
    
    # Encabezado del resumen
    summary_sheet.merge_cells('A1:D1')
    summary_sheet['A1'] = f"RESUMEN DE PRESUPUESTO - {project.name.upper()}"
    summary_sheet['A1'].font = Font(name='Arial', size=16, bold=True)
    summary_sheet['A1'].alignment = center_alignment
    summary_sheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_sheet['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    
    summary_sheet.merge_cells('A2:D2')
    summary_sheet['A2'] = f"Proyecto: {project.name}"
    summary_sheet['A2'].font = Font(name='Arial', size=12, bold=True)
    summary_sheet['A2'].alignment = center_alignment
    
    summary_sheet.merge_cells('A3:D3')
    summary_sheet['A3'] = f"Fecha de exportación: {get_colombia_time().strftime('%d/%m/%Y %H:%M')}"
    summary_sheet['A3'].font = Font(name='Arial', size=10, italic=True)
    summary_sheet['A3'].alignment = center_alignment
    
    # Encabezados del resumen
    row = 5
    headers = ['Sección', 'Descripción', 'Subtotal (COP)', 'Porcentaje (%)']
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos del resumen
    row += 1
    for section in sections_with_data:
        if section.name in total_sections:
            section_total = total_sections[section.name]
            percentage = (section_total / grand_total * 100) if grand_total > 0 else 0
            
            print(f"🔍 DEBUG Excel Export - Resumen sección {section.name}: ${section_total:,.0f} ({percentage:.1f}%)")
            
            summary_sheet.cell(row=row, column=1).value = f"Sección {section.order}"
            summary_sheet.cell(row=row, column=1).font = normal_font
            summary_sheet.cell(row=row, column=1).border = border
            
            summary_sheet.cell(row=row, column=2).value = section.name
            summary_sheet.cell(row=row, column=2).font = normal_font
            summary_sheet.cell(row=row, column=2).border = border
            
            summary_sheet.cell(row=row, column=3).value = section_total
            summary_sheet.cell(row=row, column=3).font = currency_font
            summary_sheet.cell(row=row, column=3).alignment = right_alignment
            summary_sheet.cell(row=row, column=3).border = border
            summary_sheet.cell(row=row, column=3).number_format = '"$"#,##0'
            
            summary_sheet.cell(row=row, column=4).value = percentage
            summary_sheet.cell(row=row, column=4).font = normal_font
            summary_sheet.cell(row=row, column=4).alignment = right_alignment
            summary_sheet.cell(row=row, column=4).border = border
            summary_sheet.cell(row=row, column=4).number_format = '0.00"%"'
            
            row += 1
    
    # Subtotal (costos directos)
    row += 1
    summary_sheet.merge_cells(f'A{row}:B{row}')
    summary_sheet[f'A{row}'] = "SUBTOTAL (Costos Directos)"
    summary_sheet[f'A{row}'].font = total_font
    summary_sheet[f'A{row}'].alignment = right_alignment
    summary_sheet[f'A{row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    summary_sheet[f'A{row}'].border = border
    
    summary_sheet.cell(row=row, column=3).value = grand_total
    summary_sheet.cell(row=row, column=3).font = total_font
    summary_sheet.cell(row=row, column=3).alignment = right_alignment
    summary_sheet.cell(row=row, column=3).number_format = '"$"#,##0'
    summary_sheet.cell(row=row, column=3).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    summary_sheet.cell(row=row, column=3).border = border
    
    summary_sheet.cell(row=row, column=4).value = 100.0
    summary_sheet.cell(row=row, column=4).font = total_font
    summary_sheet.cell(row=row, column=4).alignment = right_alignment
    summary_sheet.cell(row=row, column=4).number_format = '0.00"%"'
    summary_sheet.cell(row=row, column=4).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    summary_sheet.cell(row=row, column=4).border = border
    
    # Administración automática (usando porcentaje del proyecto)
    admin_percentage = float(project.administration_percentage)
    admin_auto = grand_total * (admin_percentage / 100)
    row += 1
    summary_sheet.merge_cells(f'A{row}:B{row}')
    summary_sheet[f'A{row}'] = f"Administración ({admin_percentage:.2f}%)"
    summary_sheet[f'A{row}'].font = total_font
    summary_sheet[f'A{row}'].alignment = right_alignment
    summary_sheet[f'A{row}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    summary_sheet[f'A{row}'].border = border
    
    summary_sheet.cell(row=row, column=3).value = admin_auto
    summary_sheet.cell(row=row, column=3).font = total_font
    summary_sheet.cell(row=row, column=3).alignment = right_alignment
    summary_sheet.cell(row=row, column=3).number_format = '"$"#,##0'
    summary_sheet.cell(row=row, column=3).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    summary_sheet.cell(row=row, column=3).border = border
    
    summary_sheet.cell(row=row, column=4).value = admin_percentage
    summary_sheet.cell(row=row, column=4).font = total_font
    summary_sheet.cell(row=row, column=4).alignment = right_alignment
    summary_sheet.cell(row=row, column=4).number_format = '0.00"%"'
    summary_sheet.cell(row=row, column=4).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    summary_sheet.cell(row=row, column=4).border = border
    
    # Total final
    final_total = grand_total + admin_auto
    row += 2
    summary_sheet.merge_cells(f'A{row}:B{row}')
    summary_sheet[f'A{row}'] = "TOTAL GENERAL DEL PROYECTO"
    summary_sheet[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    summary_sheet[f'A{row}'].alignment = right_alignment
    summary_sheet[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_sheet[f'A{row}'].border = border
    
    summary_sheet.cell(row=row, column=3).value = final_total
    summary_sheet.cell(row=row, column=3).font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    summary_sheet.cell(row=row, column=3).alignment = right_alignment
    summary_sheet.cell(row=row, column=3).number_format = '"$"#,##0'
    summary_sheet.cell(row=row, column=3).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_sheet.cell(row=row, column=3).border = border
    
    # Ajustar ancho de columnas del resumen
    summary_sheet.column_dimensions['A'].width = 15
    summary_sheet.column_dimensions['B'].width = 40
    summary_sheet.column_dimensions['C'].width = 20
    summary_sheet.column_dimensions['D'].width = 15
    
    # Preparar respuesta HTTP
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    print(f"🔍 DEBUG Excel Export - RESUMEN FINAL:")
    print(f"  - Total secciones procesadas: {len(sections_with_data)}")
    print(f"  - Gran total: ${grand_total:,.0f}")
    print(f"  - Administración ({admin_percentage:.2f}%): ${admin_auto:,.0f}")
    print(f"  - Total final: ${final_total:,.0f}")
    
    # Generar nombre del archivo
    project_name_clean = "".join(c for c in project.name if c.isalnum() or c in (' ', '_')).strip()
    project_name_clean = project_name_clean.replace(' ', '_')
    fecha_actual = get_colombia_time().strftime('%Y-%m-%d')
    filename = f"Presupuesto_{project_name_clean}_{fecha_actual}.xlsx"
    
    print(f"🔍 DEBUG Excel Export - Archivo generado: {filename}")
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    messages.success(request, f'✅ Presupuesto exportado exitosamente: {filename}')
    
    return response


# Función para exportar gastos diarios a Excel
@project_owner_or_jefe_required
def export_gastos_to_excel(request, project_id):
    """
    Vista para exportar gastos diarios de materiales a Excel
    Accesible para JEFE o dueño del proyecto
    Permite filtrar por día, mes o proyecto completo
    """
    project = get_object_or_404(Project, id=project_id)
    
    # Parámetros de filtrado
    tipo_filtro = request.GET.get('tipo', 'proyecto')  # 'dia', 'mes', 'proyecto'
    fecha = request.GET.get('fecha', '')  # Para filtro por día
    mes = request.GET.get('mes', '')  # Para filtro por mes (formato: YYYY-MM)
    
    print(f"🔍 DEBUG Export Gastos - Parámetros recibidos:")
    print(f"  - Tipo filtro: {tipo_filtro}")
    print(f"  - Fecha: {fecha}")
    print(f"  - Mes: {mes}")
    print(f"  - Proyecto: {project.name}")
    
    from .models import ConsumoMaterial, ProyectoMaterial
    from datetime import datetime, date
    from django.utils import timezone
    
    # Construir query base
    consumos_query = ConsumoMaterial.objects.filter(proyecto=project).select_related(
        'material', 'material__unit', 'registrado_por'
    )
    
    # Aplicar filtros según el tipo
    if tipo_filtro == 'dia' and fecha:
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
            consumos_query = consumos_query.filter(fecha_consumo=fecha_obj)
            periodo_texto = f"Día {fecha_obj.strftime('%d/%m/%Y')}"
        except ValueError:
            periodo_texto = "Día (fecha inválida)"
    elif tipo_filtro == 'mes' and mes:
        try:
            year, month = map(int, mes.split('-'))
            consumos_query = consumos_query.filter(
                fecha_consumo__year=year,
                fecha_consumo__month=month
            )
            fecha_mes = date(year, month, 1)
            
            # Nombres de meses en español
            meses_es = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
            mes_nombre = meses_es.get(month, 'Mes')
            periodo_texto = f"Mes {mes_nombre} {year}"
        except (ValueError, IndexError):
            periodo_texto = "Mes (formato inválido)"
    else:
        periodo_texto = "Proyecto completo"
    
    # Ordenar por fecha
    consumos = consumos_query.order_by('fecha_consumo', 'material__name')
    
    print(f"🔍 DEBUG Export Gastos - Total consumos encontrados: {consumos.count()}")
    
    if not consumos.exists():
        messages.warning(request, f'No se encontraron gastos para el período seleccionado: {periodo_texto}')
        return redirect('projects:project_board', project_id=project_id)
    
    # Crear workbook
    workbook = openpyxl.Workbook()
    
    # Eliminar hoja por defecto y crear nueva
    workbook.remove(workbook.active)
    ws = workbook.create_sheet("Gastos de Materiales")
    
    # ===== CONFIGURACIÓN DE ESTILOS =====
    # Fuentes
    font_title = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    font_data = Font(name='Calibri', size=11)
    font_total = Font(name='Calibri', size=12, bold=True)
    
    # Rellenos
    fill_title = PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid')
    fill_header = PatternFill(start_color='4F7942', end_color='4F7942', fill_type='solid')
    fill_total = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    
    # Bordes
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== ENCABEZADO DEL REPORTE =====
    ws.merge_cells('A1:G1')
    ws['A1'] = f"REPORTE DE GASTOS - {project.name.upper()}"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_title
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Período: {periodo_texto}"
    ws['A2'].font = Font(name='Calibri', size=12, bold=True, color='2F5233')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Generado: {get_colombia_time().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = Font(name='Calibri', size=10, color='666666')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # ===== ENCABEZADOS DE COLUMNAS =====
    headers = ['Fecha', 'Material', 'SKU', 'Cantidad', 'Unidad', 'Costo Unit.', 'Costo Total', 'Actividad', 'Responsable']
    header_row = 5
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # ===== DATOS =====
    current_row = header_row + 1
    total_general = 0
    
    for consumo in consumos:
        # Obtener costo unitario del material
        costo_unitario = float(consumo.material.unit_cost or 0)
        
        costo_total = float(consumo.cantidad_consumida) * costo_unitario
        total_general += costo_total
        
        # Datos de la fila
        row_data = [
            consumo.fecha_consumo.strftime('%d/%m/%Y'),
            consumo.material.name,
            consumo.material.sku or '',
            float(consumo.cantidad_consumida),
            consumo.material.unit.symbol,
            costo_unitario,
            costo_total,
            consumo.componente_actividad,
            consumo.responsable or (request.user.get_full_name() or request.user.username)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.font = font_data
            cell.border = border_thin
            
            # Formato específico para columnas numéricas
            if col in [4, 6, 7]:  # Cantidad, Costo Unit., Costo Total
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 1:  # Fecha
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        
        current_row += 1
    
    # ===== FILA DE TOTAL =====
    total_row = current_row + 1
    
    ws.merge_cells(f'A{total_row}:F{total_row}')
    cell_total_label = ws[f'A{total_row}']
    cell_total_label.value = "TOTAL GENERAL"
    cell_total_label.font = font_total
    cell_total_label.fill = fill_total
    cell_total_label.alignment = Alignment(horizontal='right', vertical='center')
    cell_total_label.border = border_thin
    
    cell_total_value = ws[f'G{total_row}']
    cell_total_value.value = total_general
    cell_total_value.font = font_total
    cell_total_value.fill = fill_total
    cell_total_value.number_format = '#,##0.00'
    cell_total_value.alignment = Alignment(horizontal='right', vertical='center')
    cell_total_value.border = border_thin
    
    # ===== AJUSTAR ANCHOS DE COLUMNAS =====
    column_widths = [12, 25, 15, 12, 8, 15, 15, 30, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # ===== AGREGAR RESUMEN EN SEGUNDA HOJA =====
    ws_resumen = workbook.create_sheet("Resumen")
    
    # Título del resumen
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = f"RESUMEN DE GASTOS - {project.name.upper()}"
    ws_resumen['A1'].font = font_title
    ws_resumen['A1'].fill = fill_title
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información del proyecto
    resumen_data = [
        ["Período:", periodo_texto],
        ["Total de registros:", consumos.count()],
        ["Total invertido:", f"${total_general:,.2f}"],
        ["Fecha de generación:", get_colombia_time().strftime('%d/%m/%Y %H:%M')],
        ["Generado por:", request.user.get_full_name() or request.user.username]
    ]
    
    for row, (label, value) in enumerate(resumen_data, 3):
        ws_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_resumen.cell(row=row, column=2, value=value)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 20
    ws_resumen.column_dimensions['B'].width = 30
    
    # ===== PREPARAR RESPUESTA =====
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    print(f"🔍 DEBUG Export Gastos - RESUMEN FINAL:")
    print(f"  - Total registros: {consumos.count()}")
    print(f"  - Total general: ${total_general:,.2f}")
    print(f"  - Período: {periodo_texto}")
    
    # Generar nombre del archivo
    project_name_clean = "".join(c for c in project.name if c.isalnum() or c in (' ', '_')).strip()
    project_name_clean = project_name_clean.replace(' ', '_')
    fecha_actual = get_colombia_time().strftime('%Y-%m-%d')
    
    if tipo_filtro == 'dia' and fecha:
        filename = f"Gastos_{project_name_clean}_{fecha}_{fecha_actual}.xlsx"
    elif tipo_filtro == 'mes' and mes:
        filename = f"Gastos_{project_name_clean}_{mes}_{fecha_actual}.xlsx"
    else:
        filename = f"Gastos_{project_name_clean}_Completo_{fecha_actual}.xlsx"
    
    print(f"🔍 DEBUG Export Gastos - Archivo generado: {filename}")
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    messages.success(request, f'✅ Gastos exportados exitosamente: {filename}')
    
    return response


@login_required
@project_owner_or_jefe_required
def export_comparativo_to_excel(request, project_id):
    """
    Exporta reporte comparativo de presupuesto vs gasto real a Excel
    RF: Como Jefe de obra, quiero exportar un reporte comparativo para analizar desviaciones financieras
    """
    from collections import defaultdict
    from decimal import Decimal
    from django.db.models import Sum, F
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    project = get_object_or_404(Project, id=project_id)
    
    print(f"🔍 DEBUG Export Comparativo - Iniciando para proyecto: {project.name}")
    
    # ===== OBTENER DATOS DEL PRESUPUESTO =====
    presupuesto_items = ProjectBudgetItem.objects.filter(
        project=project
    ).select_related('budget_item', 'budget_item__section').order_by(
        'budget_item__section__order', 'budget_item__order'
    )
    
    # Agrupar presupuesto por sección
    presupuesto_por_seccion = defaultdict(lambda: {
        'items': [],
        'total_presupuestado': Decimal('0'),
        'seccion_nombre': '',
        'seccion_order': 0
    })
    
    total_presupuesto_proyecto = Decimal('0')
    
    for item in presupuesto_items:
        seccion_key = item.budget_item.section.name
        seccion_data = presupuesto_por_seccion[seccion_key]
        
        seccion_data['seccion_nombre'] = item.budget_item.section.name
        seccion_data['seccion_order'] = item.budget_item.section.order
        seccion_data['items'].append({
            'descripcion': item.budget_item.description,
            'cantidad': item.quantity,
            'precio_unitario': item.unit_price,
            'total': item.total_price
        })
        seccion_data['total_presupuestado'] += item.total_price
        total_presupuesto_proyecto += item.total_price
    
    print(f"🔍 DEBUG - Total presupuesto proyecto: ${total_presupuesto_proyecto:,.2f}")
    print(f"🔍 DEBUG - Secciones de presupuesto encontradas: {len(presupuesto_por_seccion)}")
    
    # ===== OBTENER DATOS DE GASTOS REALES =====
    consumos = ConsumoMaterial.objects.filter(
        proyecto=project
    ).select_related('material', 'material__unit')
    
    # Agrupar gastos por componente/actividad
    gastos_por_componente = defaultdict(lambda: {
        'items': [],
        'total_gastado': Decimal('0')
    })
    
    total_gastos_proyecto = Decimal('0')
    
    for consumo in consumos:
        componente = consumo.componente_actividad or 'Sin especificar'
        costo_unitario = getattr(consumo.material, 'unit_cost', None) or Decimal('0')
        costo_total = consumo.cantidad_consumida * costo_unitario
        
        gastos_por_componente[componente]['items'].append({
            'material': consumo.material.name,
            'cantidad': consumo.cantidad_consumida,
            'costo_unitario': costo_unitario,
            'costo_total': costo_total,
            'fecha': consumo.fecha_consumo
        })
        gastos_por_componente[componente]['total_gastado'] += costo_total
        total_gastos_proyecto += costo_total
    
    print(f"🔍 DEBUG - Total gastos proyecto: ${total_gastos_proyecto:,.2f}")
    print(f"🔍 DEBUG - Componentes de gasto encontrados: {len(gastos_por_componente)}")
    
    # ===== CREAR WORKBOOK =====
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # ===== CONFIGURACIÓN DE ESTILOS =====
    font_title = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    font_header = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    font_subheader = Font(name='Calibri', size=11, bold=True, color='2F5233')
    font_data = Font(name='Calibri', size=11)
    font_total = Font(name='Calibri', size=12, bold=True)
    font_deviation_positive = Font(name='Calibri', size=11, bold=True, color='D32F2F')  # Rojo para sobrecostos
    font_deviation_negative = Font(name='Calibri', size=11, bold=True, color='388E3C')  # Verde para ahorros
    
    fill_title = PatternFill(start_color='2F5233', end_color='2F5233', fill_type='solid')
    fill_header = PatternFill(start_color='4F7942', end_color='4F7942', fill_type='solid')
    fill_section = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    fill_total = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
    fill_overbudget = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')  # Rojo claro
    fill_underbudget = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Verde claro
    
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: COMPARATIVO POR SECCIONES =====
    ws_comparativo = workbook.create_sheet("Comparativo Presupuesto")
    
    # Título principal
    ws_comparativo.merge_cells('A1:H1')
    ws_comparativo['A1'] = f"COMPARATIVO PRESUPUESTO VS GASTO REAL - {project.name.upper()}"
    ws_comparativo['A1'].font = font_title
    ws_comparativo['A1'].fill = fill_title
    ws_comparativo['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información del reporte
    ws_comparativo.merge_cells('A2:H2')
    ws_comparativo['A2'] = f"Generado: {get_colombia_time().strftime('%d/%m/%Y %H:%M')}"
    ws_comparativo['A2'].font = Font(name='Calibri', size=10, color='666666')
    ws_comparativo['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    headers = [
        'Sección/Componente', 'Presupuesto Proyectado', 'Gasto Real', 
        'Desviación ($)', 'Desviación (%)', 'Estado', 'Items Presupuesto', 'Items Gastados'
    ]
    
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws_comparativo.cell(row=header_row, column=col)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = header_row + 1
    
    # ===== MAPEO INTELIGENTE SECCIONES VS COMPONENTES =====
    # Crear mapeo aproximado basado en palabras clave
    mapeo_seccion_componente = {
        'cimentación': ['cimentacion', 'cimientos', 'zapata', 'fundacion'],
        'estructura': ['estructura', 'viga', 'columna', 'concreto', 'acero'],
        'muros': ['muro', 'pared', 'mamposteria', 'ladrillo', 'bloque'],
        'cubierta': ['cubierta', 'techo', 'teja', 'impermeabilizacion'],
        'pisos': ['piso', 'acabados', 'ceramica', 'baldosa'],
        'instalaciones': ['instalacion', 'electrica', 'hidrosanitaria', 'fontaneria', 'electricidad'],
        'carpinteria': ['puerta', 'ventana', 'marco', 'carpinteria'],
        'pintura': ['pintura', 'acabados'],
        'varios': ['varios', 'miscelaneos', 'otros']
    }
    
    def encontrar_seccion_para_componente(componente):
        """Encuentra la sección de presupuesto más apropiada para un componente de gasto"""
        componente_lower = componente.lower()
        for seccion_nombre, seccion_data in presupuesto_por_seccion.items():
            seccion_lower = seccion_nombre.lower()
            # Buscar coincidencia directa
            if any(palabra in componente_lower for palabra in seccion_lower.split()):
                return seccion_nombre
        
        # Buscar por mapeo de palabras clave
        for patron, palabras_clave in mapeo_seccion_componente.items():
            if any(palabra in componente_lower for palabra in palabras_clave):
                # Buscar sección que contenga el patrón
                for seccion_nombre in presupuesto_por_seccion.keys():
                    if patron in seccion_nombre.lower():
                        return seccion_nombre
        
        return None
    
    # ===== PROCESAR DATOS PARA EL COMPARATIVO =====
    comparativo_data = []
    secciones_procesadas = set()
    
    # 1. Procesar secciones del presupuesto
    for seccion_nombre, seccion_data in sorted(presupuesto_por_seccion.items(), 
                                               key=lambda x: x[1]['seccion_order']):
        presupuestado = seccion_data['total_presupuestado']
        
        # Buscar gastos relacionados con esta sección
        gastos_relacionados = Decimal('0')
        componentes_relacionados = []
        
        for componente, gasto_data in gastos_por_componente.items():
            seccion_encontrada = encontrar_seccion_para_componente(componente)
            if seccion_encontrada == seccion_nombre:
                gastos_relacionados += gasto_data['total_gastado']
                componentes_relacionados.append(componente)
        
        # Marcar componentes como procesados
        for comp in componentes_relacionados:
            secciones_procesadas.add(comp)
        
        # Calcular desviaciones
        desviacion_abs = gastos_relacionados - presupuestado
        desviacion_pct = (desviacion_abs / presupuestado * 100) if presupuestado > 0 else 0
        
        estado = "SOBRE PRESUPUESTO" if desviacion_abs > 0 else "DENTRO PRESUPUESTO" if desviacion_abs == 0 else "BAJO PRESUPUESTO"
        
        comparativo_data.append({
            'nombre': seccion_nombre,
            'presupuestado': presupuestado,
            'gastado': gastos_relacionados,
            'desviacion_abs': desviacion_abs,
            'desviacion_pct': desviacion_pct,
            'estado': estado,
            'items_presupuesto': len(seccion_data['items']),
            'items_gastados': len(componentes_relacionados),
            'tipo': 'seccion'
        })
    
    # 2. Procesar componentes de gasto sin sección asignada
    for componente, gasto_data in gastos_por_componente.items():
        if componente not in secciones_procesadas:
            gastado = gasto_data['total_gastado']
            
            comparativo_data.append({
                'nombre': f"[GASTO SIN PRESUPUESTO] {componente}",
                'presupuestado': Decimal('0'),
                'gastado': gastado,
                'desviacion_abs': gastado,
                'desviacion_pct': 100 if gastado > 0 else 0,  # 100% desviación si no estaba presupuestado
                'estado': "SIN PRESUPUESTO",
                'items_presupuesto': 0,
                'items_gastados': len(gasto_data['items']),
                'tipo': 'gasto_extra'
            })
    
    # ===== ESCRIBIR DATOS EN LA HOJA =====
    for data in comparativo_data:
        # Determinar estilo según el estado
        if data['tipo'] == 'gasto_extra':
            fill_row = fill_overbudget
            font_desviacion = font_deviation_positive
        elif data['desviacion_abs'] > 0:
            fill_row = fill_overbudget
            font_desviacion = font_deviation_positive
        elif data['desviacion_abs'] < 0:
            fill_row = fill_underbudget
            font_desviacion = font_deviation_negative
        else:
            fill_row = None
            font_desviacion = font_data
        
        row_data = [
            data['nombre'],
            float(data['presupuestado']),
            float(data['gastado']),
            float(data['desviacion_abs']),
            float(data['desviacion_pct']),
            data['estado'],
            data['items_presupuesto'],
            data['items_gastados']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_comparativo.cell(row=current_row, column=col)
            cell.value = value
            cell.border = border_thin
            
            # Aplicar formato específico
            if col in [2, 3, 4]:  # Presupuestado, Gastado, Desviación $
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 5:  # Desviación %
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='right')
                cell.font = font_desviacion
            elif col in [7, 8]:  # Contadores
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
            
            # Aplicar fondo si es necesario
            if fill_row and col <= 8:
                cell.fill = fill_row
            
            cell.font = font_data
        
        current_row += 1
    
    # ===== FILA DE TOTALES =====
    total_row = current_row + 1
    
    # Calcular totales
    total_presupuestado_final = sum(item['presupuestado'] for item in comparativo_data 
                                    if item['tipo'] == 'seccion')
    total_gastado_final = sum(item['gastado'] for item in comparativo_data)
    desviacion_total = total_gastado_final - total_presupuestado_final
    desviacion_pct_total = (desviacion_total / total_presupuestado_final * 100) if total_presupuestado_final > 0 else 0
    
    # Aplicar totales
    ws_comparativo.merge_cells(f'A{total_row}:A{total_row}')
    cell_total_label = ws_comparativo[f'A{total_row}']
    cell_total_label.value = "TOTALES GENERALES"
    cell_total_label.font = font_total
    cell_total_label.fill = fill_total
    cell_total_label.alignment = Alignment(horizontal='center', vertical='center')
    cell_total_label.border = border_thin
    
    totales_data = [
        None,  # Ya asignado arriba
        float(total_presupuestado_final),
        float(total_gastado_final),
        float(desviacion_total),
        float(desviacion_pct_total),
        "SOBRE PRESUPUESTO" if desviacion_total > 0 else "DENTRO PRESUPUESTO" if desviacion_total == 0 else "BAJO PRESUPUESTO",
        sum(item['items_presupuesto'] for item in comparativo_data),
        sum(item['items_gastados'] for item in comparativo_data)
    ]
    
    for col, value in enumerate(totales_data[1:], 2):  # Empezar desde columna 2
        cell = ws_comparativo.cell(row=total_row, column=col)
        cell.value = value
        cell.font = font_total
        cell.fill = fill_total
        cell.border = border_thin
        
        if col in [2, 3, 4]:  # Montos
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        elif col == 5:  # Porcentaje
            cell.number_format = '0.00"%"'
            cell.alignment = Alignment(horizontal='right')
        elif col in [7, 8]:  # Contadores
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(horizontal='center')
    
    # ===== AJUSTAR ANCHOS DE COLUMNAS =====
    column_widths = [35, 18, 18, 18, 15, 20, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws_comparativo.column_dimensions[get_column_letter(col)].width = width
    
    # ===== HOJA 2: RESUMEN EJECUTIVO =====
    ws_resumen = workbook.create_sheet("Resumen Ejecutivo")
    
    # Título del resumen
    ws_resumen.merge_cells('A1:D1')
    ws_resumen['A1'] = f"RESUMEN EJECUTIVO - {project.name.upper()}"
    ws_resumen['A1'].font = font_title
    ws_resumen['A1'].fill = fill_title
    ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Métricas clave
    metricas_data = [
        ["MÉTRICAS FINANCIERAS", ""],
        ["Presupuesto inicial:", f"${total_presupuestado_final:,.2f}"],
        ["Gasto real acumulado:", f"${total_gastado_final:,.2f}"],
        ["Desviación total:", f"${desviacion_total:,.2f}"],
        ["Desviación porcentual:", f"{desviacion_pct_total:.2f}%"],
        ["", ""],
        ["ANÁLISIS POR ESTADO", ""],
        ["Secciones sobre presupuesto:", len([d for d in comparativo_data if d['desviacion_abs'] > 0])],
        ["Secciones bajo presupuesto:", len([d for d in comparativo_data if d['desviacion_abs'] < 0])],
        ["Gastos sin presupuesto:", len([d for d in comparativo_data if d['tipo'] == 'gasto_extra'])],
        ["", ""],
        ["INFORMACIÓN DEL REPORTE", ""],
        ["Fecha de generación:", get_colombia_time().strftime('%d/%m/%Y %H:%M')],
        ["Generado por:", request.user.get_full_name() or request.user.username],
        ["Total de ítems presupuestados:", sum(item['items_presupuesto'] for item in comparativo_data)],
        ["Total de registros de gasto:", sum(item['items_gastados'] for item in comparativo_data)]
    ]
    
    for row, (label, value) in enumerate(metricas_data, 3):
        if label == "MÉTRICAS FINANCIERAS" or label == "ANÁLISIS POR ESTADO" or label == "INFORMACIÓN DEL REPORTE":
            # Encabezado de sección
            ws_resumen.merge_cells(f'A{row}:D{row}')
            cell = ws_resumen[f'A{row}']
            cell.value = label
            cell.font = font_subheader
            cell.fill = fill_section
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
        elif label:  # Solo si tiene contenido
            ws_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_resumen.cell(row=row, column=2, value=value)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    
    # ===== PREPARAR RESPUESTA =====
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    print(f"🔍 DEBUG Export Comparativo - RESUMEN FINAL:")
    print(f"  - Total presupuestado: ${total_presupuestado_final:,.2f}")
    print(f"  - Total gastado: ${total_gastado_final:,.2f}")
    print(f"  - Desviación: ${desviacion_total:,.2f} ({desviacion_pct_total:.2f}%)")
    print(f"  - Secciones procesadas: {len(comparativo_data)}")
    
    # Generar nombre del archivo
    project_name_clean = "".join(c for c in project.name if c.isalnum() or c in (' ', '_')).strip()
    project_name_clean = project_name_clean.replace(' ', '_')
    fecha_actual = get_colombia_time().strftime('%Y-%m-%d')
    filename = f"Comparativo_{project_name_clean}_{fecha_actual}.xlsx"
    
    print(f"🔍 DEBUG Export Comparativo - Archivo generado: {filename}")
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    messages.success(request, f'✅ Reporte comparativo exportado exitosamente: {filename}')
    
    return response