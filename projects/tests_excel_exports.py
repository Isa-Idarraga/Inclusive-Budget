"""
Pruebas automáticas para las funcionalidades de exportación a Excel.
RF: Generación de reportes Excel desde el botón desplegable

IMPORTANTE: 
- Estas pruebas NO modifican la base de datos de producción
- Usan fixtures y mocks para simular datos
- Cada prueba es independiente y aislada
- Se ejecutan en una base de datos de prueba (SQLite)

Autor: Sistema de pruebas automáticas
Fecha: 2025-10-18
"""

from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from decimal import Decimal
from datetime import date, timedelta
import openpyxl
from io import BytesIO

# Importar modelos necesarios
from projects.models import (
    Project, 
    BudgetSection, 
    BudgetItem, 
    ProjectBudgetItem,
    ConsumoMaterial,
    EntradaMaterial
)
from catalog.models import Material, Unit

User = get_user_model()


class ExcelExportTestCase(TestCase):
    """Clase base para pruebas de exportación Excel"""
    
    def setUp(self):
        """Configuración inicial para cada prueba"""
        # Crear usuarios con diferentes roles
        self.jefe = User.objects.create_user(
            username='jefe_test',
            email='jefe@test.com',
            password='testpass123',
            role=User.JEFE,
            first_name='Jefe',
            last_name='Test'
        )
        
        self.constructor = User.objects.create_user(
            username='constructor_test',
            email='constructor@test.com',
            password='testpass123',
            role=User.CONSTRUCTOR,
            first_name='Constructor',
            last_name='Test'
        )
        
        self.comercial = User.objects.create_user(
            username='comercial_test',
            email='comercial@test.com',
            password='testpass123',
            role=User.COMERCIAL,
            first_name='Comercial',
            last_name='Test'
        )
        
        # Crear unidades
        self.unit_m2 = Unit.objects.create(
            name='Metro cuadrado',
            symbol='m²'
        )
        
        self.unit_ml = Unit.objects.create(
            name='Metro lineal',
            symbol='ml'
        )
        
        self.unit_kg = Unit.objects.create(
            name='Kilogramo',
            symbol='kg'
        )
        
        # Crear materiales
        self.material_cemento = Material.objects.create(
            sku='CEM-001',
            name='Cemento Portland',
            unit=self.unit_kg,
            unit_cost=Decimal('45000'),
            category='CEMENTOS',
            stock=Decimal('100.000'),
            presentation_qty=Decimal('50.000')
        )
        
        self.material_arena = Material.objects.create(
            sku='ARE-001',
            name='Arena',
            unit=self.unit_kg,
            unit_cost=Decimal('15000'),
            category='AGREGADOS',
            stock=Decimal('500.000'),
            presentation_qty=Decimal('1.000')
        )
        
        self.material_ladrillo = Material.objects.create(
            sku='LAD-001',
            name='Ladrillo',
            unit=self.unit_kg,
            unit_cost=Decimal('1200'),
            category='LADRILLOS',
            stock=Decimal('1000.000'),
            presentation_qty=Decimal('1.000')
        )
        
        # Crear proyecto de prueba
        self.project = Project.objects.create(
            name='Proyecto Test Excel',
            description='Proyecto para pruebas de exportación',
            location_address='Medellín, Colombia',
            presupuesto=Decimal('50000000.00'),
            built_area=Decimal('100.00'),
            exterior_area=Decimal('50.00'),
            columns_count=10,
            walls_area=Decimal('200.00'),
            windows_area=Decimal('20.00'),
            doors_count=5,
            doors_height=Decimal('2.10'),
            ubicacion_proyecto='Medellin',
            area_construida_total=Decimal('150.00'),
            numero_pisos='2',
            creado_por=self.constructor
        )
        
        # Crear secciones del presupuesto
        self.section_preliminares = BudgetSection.objects.create(
            name='PRELIMINARES',
            description='Trabajos preliminares',
            order=1
        )
        
        self.section_cimientos = BudgetSection.objects.create(
            name='CIMIENTOS',
            description='Trabajos de cimentación',
            order=2
        )
        
        # Crear ítems de presupuesto
        self.item_limpieza = BudgetItem.objects.create(
            section=self.section_preliminares,
            code='P-001',
            name='Limpieza del terreno',
            description='Limpieza y nivelación',
            unit=self.unit_m2,
            unit_price=Decimal('25000.00'),
            order=1,
            is_active=True
        )
        
        self.item_excavacion = BudgetItem.objects.create(
            section=self.section_cimientos,
            code='C-001',
            name='Excavación manual',
            description='Excavación para cimientos',
            unit=self.unit_m2,
            unit_price=Decimal('35000.00'),
            order=1,
            is_active=True
        )
        
        # Asignar ítems al proyecto
        self.project_item_limpieza = ProjectBudgetItem.objects.create(
            project=self.project,
            budget_item=self.item_limpieza,
            quantity=Decimal('100.00')
        )
        
        self.project_item_excavacion = ProjectBudgetItem.objects.create(
            project=self.project,
            budget_item=self.item_excavacion,
            quantity=Decimal('50.00')
        )
        
        # Crear entradas de materiales
        self.entrada_cemento = EntradaMaterial.objects.create(
            material=self.material_cemento,
            proyecto=self.project,
            cantidad=Decimal('100.00'),
            fecha_entrada=date.today() - timedelta(days=10)
        )
        
        self.entrada_arena = EntradaMaterial.objects.create(
            material=self.material_arena,
            proyecto=self.project,
            cantidad=Decimal('200.00'),
            fecha_entrada=date.today() - timedelta(days=8)
        )
        
        # Crear consumos de materiales
        self.consumo_cemento = ConsumoMaterial.objects.create(
            material=self.material_cemento,
            proyecto=self.project,
            cantidad_consumida=Decimal('50.00'),
            fecha_consumo=date.today() - timedelta(days=5),
            responsable=self.constructor
        )
        
        self.consumo_arena = ConsumoMaterial.objects.create(
            material=self.material_arena,
            proyecto=self.project,
            cantidad_consumida=Decimal('100.00'),
            fecha_consumo=date.today() - timedelta(days=3),
            responsable=self.constructor
        )
        
        # Cliente para hacer requests
        self.client = Client()


class ExportBudgetToExcelTests(ExcelExportTestCase):
    """Pruebas para export_budget_to_excel"""
    
    def test_export_budget_requires_authentication(self):
        """Verifica que se requiere autenticación"""
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        # Debe redirigir al login
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)
    
    def test_export_budget_requires_jefe_role(self):
        """Verifica que solo usuarios JEFE pueden exportar"""
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        
        # Probar con comercial (no debe tener acceso)
        self.client.login(username='comercial_test', password='testpass123')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)  # Redirigido
        
        # Probar con jefe (debe tener acceso)
        self.client.login(username='jefe_test', password='testpass123')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
    
    def test_export_budget_returns_excel_file(self):
        """Verifica que la respuesta sea un archivo Excel válido"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        # Verificar headers
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('Presupuesto_', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
    
    def test_export_budget_excel_structure(self):
        """Verifica la estructura del archivo Excel generado"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        # Cargar el Excel generado
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        
        # Verificar que existe hoja de RESUMEN
        self.assertIn('RESUMEN', workbook.sheetnames)
        
        # Verificar que existen hojas de secciones
        sheet_names = workbook.sheetnames
        self.assertTrue(any('PRELIMINARES' in name for name in sheet_names))
        self.assertTrue(any('CIMIENTOS' in name for name in sheet_names))
        
        # Verificar contenido de hoja RESUMEN
        resumen = workbook['RESUMEN']
        titulo = resumen['A1'].value
        self.assertIn('RESUMEN DE PRESUPUESTO', titulo.upper())
        self.assertIn(self.project.name.upper(), titulo.upper())
    
    def test_export_budget_contains_project_info(self):
        """Verifica que el Excel contenga información del proyecto"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        
        # Revisar una hoja de sección
        for sheet_name in workbook.sheetnames:
            if 'RESUMEN' not in sheet_name:
                sheet = workbook[sheet_name]
                # Verificar que tenga contenido
                self.assertIsNotNone(sheet['A1'].value)
                break
    
    def test_export_budget_without_items_shows_message(self):
        """Verifica comportamiento cuando no hay ítems configurados"""
        # Crear proyecto sin ítems
        project_sin_items = Project.objects.create(
            name='Proyecto Sin Items',
            description='Test',
            location_address='Test',
            presupuesto=Decimal('1000000.00'),
            built_area=Decimal('50.00'),
            exterior_area=Decimal('25.00'),
            columns_count=5,
            walls_area=Decimal('100.00'),
            windows_area=Decimal('10.00'),
            doors_count=3,
            doors_height=Decimal('2.10'),
            ubicacion_proyecto='Medellin',
            area_construida_total=Decimal('75.00'),
            numero_pisos='1',
            creado_por=self.constructor
        )
        
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_budget_to_excel', args=[project_sin_items.id])
        response = self.client.get(url)
        
        # Debe redirigir con mensaje
        self.assertEqual(response.status_code, 302)


class ExportGastosToExcelTests(ExcelExportTestCase):
    """Pruebas para export_gastos_to_excel"""
    
    def test_export_gastos_requires_authentication(self):
        """Verifica que se requiere autenticación"""
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)
    
    def test_export_gastos_jefe_can_access(self):
        """Verifica que JEFE puede exportar gastos"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response = self.client.get(url, {'tipo': 'proyecto'})
        
        self.assertEqual(response.status_code, 200)
    
    def test_export_gastos_constructor_can_access(self):
        """Verifica que el constructor puede exportar gastos"""
        self.client.login(username='constructor_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response = self.client.get(url, {'tipo': 'proyecto'})
        
        self.assertEqual(response.status_code, 200)
    
    def test_export_gastos_proyecto_completo(self):
        """Verifica exportación de gastos del proyecto completo"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response = self.client.get(url, {'tipo': 'proyecto'})
        
        # Verificar respuesta
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('Gastos_', response['Content-Disposition'])
        
        # Verificar estructura del Excel
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        self.assertIn('Gastos de Materiales', workbook.sheetnames)
        self.assertIn('Resumen', workbook.sheetnames)
    
    def test_export_gastos_por_mes(self):
        """Verifica exportación de gastos filtrados por mes"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        mes = date.today().strftime('%Y-%m')
        response = self.client.get(url, {'tipo': 'mes', 'mes': mes})
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('.xlsx', response['Content-Disposition'])
        
        # Verificar que el nombre incluye el mes
        self.assertIn(mes, response['Content-Disposition'])
    
    def test_export_gastos_por_dia(self):
        """Verifica exportación de gastos filtrados por día"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        fecha = date.today().strftime('%Y-%m-%d')
        response = self.client.get(url, {'tipo': 'dia', 'fecha': fecha})
        
        self.assertEqual(response.status_code, 200)
        self.assertIn('.xlsx', response['Content-Disposition'])
    
    def test_export_gastos_excel_contains_data(self):
        """Verifica que el Excel de gastos contenga datos"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response = self.client.get(url, {'tipo': 'proyecto'})
        
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        ws = workbook['Gastos de Materiales']
        
        # Verificar título
        titulo = ws['A1'].value
        self.assertIn('REPORTE DE GASTOS', titulo.upper())
        self.assertIn(self.project.name.upper(), titulo.upper())
        
        # Verificar que hay encabezados
        self.assertIsNotNone(ws['A4'].value)  # Debe haber encabezados en fila 4
    
    def test_export_gastos_without_data_shows_warning(self):
        """Verifica comportamiento cuando no hay gastos"""
        # Crear proyecto sin consumos
        project_sin_gastos = Project.objects.create(
            name='Proyecto Sin Gastos',
            description='Test',
            location_address='Test',
            presupuesto=Decimal('1000000.00'),
            built_area=Decimal('50.00'),
            exterior_area=Decimal('25.00'),
            columns_count=5,
            walls_area=Decimal('100.00'),
            windows_area=Decimal('10.00'),
            doors_count=3,
            doors_height=Decimal('2.10'),
            ubicacion_proyecto='Medellin',
            area_construida_total=Decimal('75.00'),
            numero_pisos='1',
            creado_por=self.constructor
        )
        
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[project_sin_gastos.id])
        response = self.client.get(url, {'tipo': 'proyecto'})
        
        # Debe redirigir con warning
        self.assertEqual(response.status_code, 302)


class ExportComparativoToExcelTests(ExcelExportTestCase):
    """Pruebas para export_comparativo_to_excel"""
    
    def test_export_comparativo_requires_authentication(self):
        """Verifica que se requiere autenticación"""
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.url)
    
    def test_export_comparativo_jefe_can_access(self):
        """Verifica que JEFE puede exportar comparativo"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
    
    def test_export_comparativo_constructor_can_access(self):
        """Verifica que el constructor puede exportar comparativo"""
        self.client.login(username='constructor_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, 200)
    
    def test_export_comparativo_returns_excel_file(self):
        """Verifica que la respuesta sea un archivo Excel válido"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        # Verificar headers
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('Comparativo_', response['Content-Disposition'])
    
    def test_export_comparativo_excel_structure(self):
        """Verifica la estructura del Excel comparativo"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        # Cargar el Excel generado
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        
        # Verificar hojas esperadas
        self.assertIn('Comparativo Presupuesto', workbook.sheetnames)
        self.assertIn('Resumen Ejecutivo', workbook.sheetnames)
    
    def test_export_comparativo_contains_comparison_data(self):
        """Verifica que el Excel contenga datos de comparación"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        ws_comparativo = workbook['Comparativo Presupuesto']
        
        # Verificar título
        titulo = ws_comparativo['A1'].value
        self.assertIn('COMPARATIVO', titulo.upper())
        self.assertIn('PRESUPUESTO', titulo.upper())
        self.assertIn('GASTO REAL', titulo.upper())
        
        # Verificar encabezados (fila 4)
        encabezados_esperados = ['Presupuesto Proyectado', 'Gasto Real', 'Desviación']
        fila_4_values = [ws_comparativo.cell(4, col).value for col in range(1, 9)]
        fila_4_text = ' '.join(str(v) for v in fila_4_values if v)
        
        for encabezado in encabezados_esperados:
            self.assertIn(encabezado, fila_4_text)
    
    def test_export_comparativo_resumen_ejecutivo(self):
        """Verifica el contenido del resumen ejecutivo"""
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response = self.client.get(url)
        
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        ws_resumen = workbook['Resumen Ejecutivo']
        
        # Verificar título
        titulo = ws_resumen['A1'].value
        self.assertIn('RESUMEN EJECUTIVO', titulo.upper())
        
        # Verificar que contiene métricas clave
        contenido = []
        for row in range(1, 20):
            for col in range(1, 5):
                val = ws_resumen.cell(row, col).value
                if val:
                    contenido.append(str(val))
        
        contenido_texto = ' '.join(contenido)
        self.assertIn('MÉTRICAS FINANCIERAS', contenido_texto)


class ExcelExportIntegrationTests(ExcelExportTestCase):
    """Pruebas de integración para el flujo completo de exportación"""
    
    def test_export_all_types_sequential(self):
        """Verifica que se puedan exportar los 3 tipos de reportes secuencialmente"""
        self.client.login(username='jefe_test', password='testpass123')
        
        # 1. Exportar presupuesto
        url_budget = reverse('projects:export_budget_to_excel', args=[self.project.id])
        response_budget = self.client.get(url_budget)
        self.assertEqual(response_budget.status_code, 200)
        
        # 2. Exportar gastos
        url_gastos = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        response_gastos = self.client.get(url_gastos, {'tipo': 'proyecto'})
        self.assertEqual(response_gastos.status_code, 200)
        
        # 3. Exportar comparativo
        url_comparativo = reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        response_comparativo = self.client.get(url_comparativo)
        self.assertEqual(response_comparativo.status_code, 200)
    
    def test_excel_files_are_valid_openpyxl(self):
        """Verifica que todos los Excel generados sean válidos para openpyxl"""
        self.client.login(username='jefe_test', password='testpass123')
        
        urls = [
            reverse('projects:export_budget_to_excel', args=[self.project.id]),
            reverse('projects:export_gastos_to_excel', args=[self.project.id]),
            reverse('projects:export_comparativo_to_excel', args=[self.project.id])
        ]
        
        for url in urls:
            params = {'tipo': 'proyecto'} if 'gastos' in url else {}
            response = self.client.get(url, params)
            
            # Intentar cargar con openpyxl
            try:
                workbook = openpyxl.load_workbook(BytesIO(response.content))
                self.assertIsNotNone(workbook)
                self.assertGreater(len(workbook.sheetnames), 0)
            except Exception as e:
                self.fail(f"Error al cargar Excel de {url}: {str(e)}")
    
    def test_export_with_project_id_not_found(self):
        """Verifica comportamiento con ID de proyecto inválido"""
        self.client.login(username='jefe_test', password='testpass123')
        
        url = reverse('projects:export_budget_to_excel', args=[99999])
        response = self.client.get(url)
        
        # Debe retornar 404
        self.assertEqual(response.status_code, 404)


class ExcelExportPerformanceTests(ExcelExportTestCase):
    """Pruebas de rendimiento para las exportaciones"""
    
    def test_export_budget_performance(self):
        """Verifica que la exportación de presupuesto sea rápida"""
        import time
        
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_budget_to_excel', args=[self.project.id])
        
        start_time = time.time()
        response = self.client.get(url)
        end_time = time.time()
        
        elapsed_time = end_time - start_time
        
        self.assertEqual(response.status_code, 200)
        # La exportación debe tomar menos de 5 segundos
        self.assertLess(elapsed_time, 5.0, 
                       f"La exportación tomó {elapsed_time:.2f} segundos")
    
    def test_export_gastos_performance(self):
        """Verifica que la exportación de gastos sea rápida"""
        import time
        
        self.client.login(username='jefe_test', password='testpass123')
        url = reverse('projects:export_gastos_to_excel', args=[self.project.id])
        
        start_time = time.time()
        response = self.client.get(url, {'tipo': 'proyecto'})
        end_time = time.time()
        
        elapsed_time = end_time - start_time
        
        self.assertEqual(response.status_code, 200)
        self.assertLess(elapsed_time, 5.0)


# ============================================================
# UTILIDADES PARA EJECUTAR LAS PRUEBAS
# ============================================================

def run_excel_export_tests():
    """
    Función auxiliar para ejecutar solo estas pruebas.
    
    Uso desde la línea de comandos:
    python manage.py test projects.tests_excel_exports --settings=core.settings_test
    """
    pass


if __name__ == '__main__':
    import django
    from django.conf import settings
    from django.test.utils import get_runner
    
    if not settings.configured:
        settings.configure(
            DATABASES={
                'default': {
                    'ENGINE': 'django.db.backends.sqlite3',
                    'NAME': ':memory:',
                }
            },
            INSTALLED_APPS=[
                'django.contrib.contenttypes',
                'django.contrib.auth',
                'projects',
                'catalog',
            ],
            SECRET_KEY='test-secret-key',
        )
    
    django.setup()
    TestRunner = get_runner(settings)
    test_runner = TestRunner()
    failures = test_runner.run_tests(['projects.tests_excel_exports'])
